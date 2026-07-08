import argparse
import csv
import json
import re
import sys
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Optional

import numpy as np
from PIL import Image, ImageDraw


PYWRAPPER_DIR = Path(__file__).resolve().parents[1] / "resource" / "pywrapper"
if str(PYWRAPPER_DIR) not in sys.path:
    sys.path.insert(0, str(PYWRAPPER_DIR))

import api_server


DEFAULT_FOCUS_POINT = "PointF(299.2863464355469, 285.9410705566406)"
AFTER_STRATEGY_LABELS = {
    "ROI2峰值帧": "roi2_peak",
    "最后一帧": "last",
}
AFTER_STRATEGY_VALUES = {value: label for label, value in AFTER_STRATEGY_LABELS.items()}
PREVIEW_MAX_SIZE = (760, 460)
PREVIEW_IMAGE_ANCHOR = "nw"
ROI_STATS_PANEL_MIN_WIDTH = 840
ROI_STATS_CARD_COLUMNS = 2
ROI_STATS_VALUE_COLUMNS = 2
ROI_STATS_FONT_SIZE = 8
ROI_STATS_ROW_PADDING = 0
ROI_NAMES = ("ROI1", "ROI2", "ROI3", "ROI4")
ROI_RECT_FIELDS = ("x", "y", "width", "height")
GUI_SETTINGS_KEY = "hem_roi2_batch_analyzer"
ROI_RECT_OVERRIDES_KEY = "roi_rect_overrides"
ROI_DEFINITIONS_KEY = "roi_definitions"
HIGHLIGHT_RULE_KEY = "highlight_rule"
EXCEL_OUTPUT_DIR_KEY = "excel_output_dir"
ROI_SHAPE_VALUES = ("rectangle", "ellipse")
ROI_SHAPE_LABELS = {"矩形": "rectangle", "椭圆": "ellipse"}
ROI_SHAPE_DISPLAY = {value: label for label, value in ROI_SHAPE_LABELS.items()}
HIGHLIGHT_MODE_LABELS = {"固定灰度值": "fixed_gray", "基线倍数": "baseline_multiplier"}
HIGHLIGHT_MODE_DISPLAY = {value: label for label, value in HIGHLIGHT_MODE_LABELS.items()}
DEFAULT_FIXED_GRAY_THRESHOLD = 93.0
DEFAULT_EXCEL_OUTPUT_DIR = Path("doc") / "tasks" / "hem-roi2-batch-analyzer" / "excel"
ROI1_COLOR = api_server.ROI1_MARKER_COLOR
ROI2_COLOR = api_server.ROI2_MARKER_COLOR
ROI3_COLOR = api_server.ROI3_MARKER_COLOR
ROI4_COLOR = api_server.ROI4_MARKER_COLOR
FOCUS_COLOR = api_server.FOCUS_MARKER_COLOR
ROI2_DEFAULT_PARAMS = {"left": 40, "right": 40, "top": 50, "bottom": 30}
ROI3_DEFAULT_PARAMS = {"left": 30, "right": 30, "top": 50, "bottom": 100}
HEM_THRESHOLD_MEAN_MULTIPLIER = 1.15
HEM_Z_SCORE_THRESHOLD = 3.0

ROI_STAT_DISPLAY_FIELDS = [
    ("rect", "位置"),
    ("size", "尺寸"),
    ("area", "面积"),
    ("mean", "均灰"),
    ("density", "灰密"),
    ("std", "标准差"),
    ("median", "中位"),
    ("median_abs_deviation", "中偏差"),
    ("p10", "P10"),
    ("p90", "P90"),
    ("threshold", "阈值"),
    ("highlight_count", "高亮数"),
    ("highlight_ratio", "高亮比"),
    ("highlight_std", "高亮差"),
    ("hem_z_area", "HEM面"),
    ("mean_delta", "均差"),
    ("mean_delta_pct", "均变率"),
    ("highlight_area_delta", "亮面差"),
]


def roi_stat_card_grid_position(index: int) -> tuple[int, int]:
    return index // ROI_STATS_CARD_COLUMNS, index % ROI_STATS_CARD_COLUMNS


def roi_stat_value_grid_position(index: int) -> tuple[int, int]:
    return index // ROI_STATS_VALUE_COLUMNS, index % ROI_STATS_VALUE_COLUMNS


def _format_roi_stat_display_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    text = str(value)
    if "," in text or "x" in text or "×" in text:
        return text
    try:
        numeric = float(text)
    except ValueError:
        return text
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.3f}".rstrip("0").rstrip(".")


SUMMARY_FIELDS = [
    "sequence",
    "status",
    "error",
    "frame_count",
    "focus_anchor",
    "offset_anchor",
    "roi2_rect",
    "before_frame_index",
    "before_frame",
    "before_mean",
    "after_frame_index",
    "after_frame",
    "after_mean",
    "roi2_diff",
    "difference_threshold",
    "roi2_color",
    "after_strategy",
]

FRAME_FIELDS = [
    "sequence",
    "frame_index",
    "frame",
    "roi2_mean",
    "roi2_diff_from_before",
]


@dataclass(frozen=True)
class AnalyzerConfig:
    root_dir: Path
    output_csv: Path
    per_frame_csv: Optional[Path]
    settings_path: Optional[Path]
    focus_point: Any
    focus_points_csv: Optional[Path]
    provider_depth_mm: Optional[float]
    focus_y_offset_mm: float
    roi2_extension_params: dict
    difference_threshold: float
    before_frame_index: int
    after_strategy: str
    include_selected_debug: bool
    max_sequences: Optional[int]
    roi3_extension_params: dict = field(default_factory=lambda: dict(ROI3_DEFAULT_PARAMS))
    roi4_rect: Optional[tuple[int, int, int, int]] = None
    roi4_bottom_region_ratio: Optional[float] = None
    roi_rect_overrides: dict[str, tuple[int, int, int, int]] = field(default_factory=dict)
    roi_definitions: dict[str, dict[str, Any]] = field(default_factory=dict)
    highlight_rule: dict[str, Any] = field(default_factory=dict)
    excel_output_dir: Path = DEFAULT_EXCEL_OUTPUT_DIR


@dataclass(frozen=True)
class GuiState:
    root_dir: str
    output_csv: str
    per_frame_csv: str
    settings_path: str
    focus_point: str
    focus_points_csv: str
    provider_depth_mm: str
    focus_y_offset_mm: str
    roi2_left: str
    roi2_right: str
    roi2_top: str
    roi2_bottom: str
    difference_threshold: str
    before_frame_index: str
    after_strategy: str
    include_selected_debug: bool
    max_sequences: str
    roi3_left: str = ""
    roi3_right: str = ""
    roi3_top: str = ""
    roi3_bottom: str = ""
    roi4_x: str = ""
    roi4_y: str = ""
    roi4_width: str = ""
    roi4_height: str = ""
    roi4_bottom_region_ratio: str = ""
    roi_rect_inputs: dict[str, dict[str, str]] = field(default_factory=dict)
    roi_shape_inputs: dict[str, str] = field(default_factory=dict)
    highlight_mode: str = "baseline_multiplier"
    highlight_fixed_gray: str = str(DEFAULT_FIXED_GRAY_THRESHOLD)
    highlight_baseline_multiplier: str = str(HEM_THRESHOLD_MEAN_MULTIPLIER)
    excel_output_dir: str = str(DEFAULT_EXCEL_OUTPUT_DIR)


def _fmt_float(value: Optional[float]) -> str:
    if value is None:
        return ""
    return f"{float(value):.6f}"


def _fmt_point(value: Optional[tuple[int, int]]) -> str:
    if value is None:
        return ""
    return f"{int(value[0])},{int(value[1])}"


def _fmt_rect(value: Optional[tuple[int, int, int, int]]) -> str:
    if value is None:
        return ""
    return ",".join(str(int(v)) for v in value)


def _parse_focus_point_value(value: Any) -> Optional[tuple[int, int]]:
    if value is None or value == "":
        return None
    if isinstance(value, (list, tuple)):
        if len(value) != 2:
            raise ValueError(f"focus_point list must contain exactly two values, got {value!r}")
        return int(value[0]), int(value[1])
    return api_server.parse_focus_point(value)


def _load_settings(path: Optional[Path]) -> dict:
    if path is None:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"settings file not found: {path}")
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError(f"settings file must contain a JSON object: {path}")
    return payload


def _parse_roi_rect_payload(value: Any, label: str) -> tuple[int, int, int, int]:
    if not isinstance(value, dict):
        raise ValueError(f"{label} must be an object with x/y/width/height")
    missing = [field for field in ROI_RECT_FIELDS if field not in value]
    if missing:
        raise ValueError(f"{label} missing fields: {', '.join(missing)}")
    x = int(value["x"])
    y = int(value["y"])
    width = int(value["width"])
    height = int(value["height"])
    if x < 0 or y < 0:
        raise ValueError(f"{label} x/y must be >= 0")
    if width <= 0 or height <= 0:
        raise ValueError(f"{label} width/height must be > 0")
    return x, y, x + width, y + height


def _settings_roi_rect_overrides(settings: dict) -> dict[str, tuple[int, int, int, int]]:
    tool_settings = settings.get(GUI_SETTINGS_KEY)
    if tool_settings is None:
        return {}
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    payload = tool_settings.get(ROI_RECT_OVERRIDES_KEY, {})
    if payload is None:
        return {}
    if not isinstance(payload, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY}.{ROI_RECT_OVERRIDES_KEY} must be an object")
    overrides: dict[str, tuple[int, int, int, int]] = {}
    for roi_name in ROI_NAMES:
        if roi_name in payload:
            overrides[roi_name] = _parse_roi_rect_payload(payload[roi_name], f"{GUI_SETTINGS_KEY}.{ROI_RECT_OVERRIDES_KEY}.{roi_name}")
    return overrides


def _normalize_roi_shape(value: Any, label: str) -> str:
    shape = ROI_SHAPE_LABELS.get(str(value), str(value)).strip()
    if shape not in ROI_SHAPE_VALUES:
        raise ValueError(f"{label} shape must be rectangle or ellipse")
    return shape


def _parse_roi_definition_payload(value: Any, label: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise ValueError(f"{label} must be an object with shape/x/y/width/height")
    return {
        "shape": _normalize_roi_shape(value.get("shape", "rectangle"), label),
        "rect": _parse_roi_rect_payload(value, label),
    }


def _settings_roi_definitions(settings: dict) -> dict[str, dict[str, Any]]:
    tool_settings = settings.get(GUI_SETTINGS_KEY)
    if tool_settings is None:
        return {
            roi_name: {"shape": "rectangle", "rect": rect}
            for roi_name, rect in _settings_roi_rect_overrides(settings).items()
        }
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    payload = tool_settings.get(ROI_DEFINITIONS_KEY)
    if payload is None:
        return {
            roi_name: {"shape": "rectangle", "rect": rect}
            for roi_name, rect in _settings_roi_rect_overrides(settings).items()
        }
    if not isinstance(payload, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY}.{ROI_DEFINITIONS_KEY} must be an object")
    definitions: dict[str, dict[str, Any]] = {}
    for roi_name in ROI_NAMES:
        if roi_name in payload:
            definitions[roi_name] = _parse_roi_definition_payload(
                payload[roi_name],
                f"{GUI_SETTINGS_KEY}.{ROI_DEFINITIONS_KEY}.{roi_name}",
            )
    return definitions


def _rect_to_input_values(rect: Optional[tuple[int, int, int, int]]) -> dict[str, str]:
    if rect is None:
        return {field: "" for field in ROI_RECT_FIELDS}
    x1, y1, x2, y2 = [int(v) for v in rect]
    return {
        "x": str(x1),
        "y": str(y1),
        "width": str(x2 - x1),
        "height": str(y2 - y1),
    }


def _roi_rect_from_input_values(values: dict[str, str], roi_name: str) -> Optional[tuple[int, int, int, int]]:
    fields = [str(values.get(field, "")).strip() for field in ROI_RECT_FIELDS]
    if all(not field for field in fields):
        return None
    if any(not field for field in fields):
        raise ValueError(f"{roi_name}区域需要同时填写X、Y、宽、高，或全部留空")
    x = _required_int(fields[0], f"{roi_name} X")
    y = _required_int(fields[1], f"{roi_name} Y")
    width = _required_int(fields[2], f"{roi_name}宽")
    height = _required_int(fields[3], f"{roi_name}高")
    if x < 0 or y < 0:
        raise ValueError(f"{roi_name} X/Y必须>=0")
    if width <= 0 or height <= 0:
        raise ValueError(f"{roi_name}宽/高必须>0")
    return x, y, x + width, y + height


def _roi_rect_overrides_from_inputs(inputs: dict[str, dict[str, str]]) -> dict[str, tuple[int, int, int, int]]:
    overrides: dict[str, tuple[int, int, int, int]] = {}
    for roi_name in ROI_NAMES:
        rect = _roi_rect_from_input_values(inputs.get(roi_name, {}), roi_name)
        if rect is not None:
            overrides[roi_name] = rect
    return overrides


def _roi_definitions_from_inputs(
    rect_inputs: dict[str, dict[str, str]],
    shape_inputs: dict[str, str],
) -> dict[str, dict[str, Any]]:
    definitions: dict[str, dict[str, Any]] = {}
    for roi_name in ROI_NAMES:
        rect = _roi_rect_from_input_values(rect_inputs.get(roi_name, {}), roi_name)
        if rect is None:
            continue
        shape = _normalize_roi_shape(shape_inputs.get(roi_name, "rectangle"), roi_name)
        definitions[roi_name] = {"shape": shape, "rect": rect}
    return definitions


def _roi_definitions_from_inputs_or_preview(
    rect_inputs: dict[str, dict[str, str]],
    shape_inputs: dict[str, str],
    preview_meta: Optional[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    definitions: dict[str, dict[str, Any]] = {}
    preview_meta = preview_meta or {}
    for roi_name in ROI_NAMES:
        rect = _roi_rect_from_input_values(rect_inputs.get(roi_name, {}), roi_name)
        if rect is None:
            rect = preview_meta.get(f"{roi_name.lower()}_rect")
        if rect is None:
            continue
        shape = _normalize_roi_shape(shape_inputs.get(roi_name, "rectangle"), roi_name)
        definitions[roi_name] = {"shape": shape, "rect": tuple(int(v) for v in rect)}
    return definitions


def _roi_rect_overrides_from_definitions(definitions: dict[str, dict[str, Any]]) -> dict[str, tuple[int, int, int, int]]:
    return {
        roi_name: tuple(definition["rect"])
        for roi_name, definition in definitions.items()
        if roi_name in ROI_NAMES and definition.get("rect") is not None
    }


def _rect_to_settings_payload(rect: tuple[int, int, int, int]) -> dict[str, int]:
    x1, y1, x2, y2 = [int(v) for v in rect]
    return {"x": x1, "y": y1, "width": x2 - x1, "height": y2 - y1}


def _validate_roi_rect_for_image(
    rect: tuple[int, int, int, int],
    width: int,
    height: int,
    label: str,
) -> tuple[int, int, int, int]:
    x1, y1, x2, y2 = [int(v) for v in rect]
    if x1 < 0 or y1 < 0:
        raise ValueError(f"{label} X/Y must be >= 0")
    if x2 <= x1 or y2 <= y1:
        raise ValueError(f"{label} width/height must be > 0")
    if x2 > width or y2 > height:
        raise ValueError(f"{label} rectangle is outside image bounds: rect={rect}, image_size={(width, height)}")
    return x1, y1, x2, y2


def save_roi_rect_overrides(settings_path: Path, overrides: dict[str, Optional[tuple[int, int, int, int]]]) -> None:
    settings = _load_settings(settings_path)
    tool_settings = settings.setdefault(GUI_SETTINGS_KEY, {})
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    persisted = {
        roi_name: _rect_to_settings_payload(rect)
        for roi_name, rect in overrides.items()
        if roi_name in ROI_NAMES and rect is not None
    }
    tool_settings[ROI_RECT_OVERRIDES_KEY] = persisted
    settings_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _default_highlight_rule() -> dict[str, Any]:
    return {
        "mode": "baseline_multiplier",
        "fixed_gray": DEFAULT_FIXED_GRAY_THRESHOLD,
        "baseline_multiplier": HEM_THRESHOLD_MEAN_MULTIPLIER,
    }


def _normalize_highlight_rule(rule: Optional[dict[str, Any]]) -> dict[str, Any]:
    if rule is None:
        return _default_highlight_rule()
    if not isinstance(rule, dict):
        raise ValueError("highlight_rule must be an object")
    mode = HIGHLIGHT_MODE_LABELS.get(str(rule.get("mode", "baseline_multiplier")), str(rule.get("mode", "baseline_multiplier"))).strip()
    if mode not in {"fixed_gray", "baseline_multiplier"}:
        raise ValueError("highlight_rule.mode must be fixed_gray or baseline_multiplier")
    fixed_gray = float(rule.get("fixed_gray", DEFAULT_FIXED_GRAY_THRESHOLD))
    baseline_multiplier = float(rule.get("baseline_multiplier", HEM_THRESHOLD_MEAN_MULTIPLIER))
    if fixed_gray < 0 or fixed_gray > 255:
        raise ValueError("highlight_rule.fixed_gray must be in 0..255")
    if baseline_multiplier <= 0:
        raise ValueError("highlight_rule.baseline_multiplier must be > 0")
    return {"mode": mode, "fixed_gray": fixed_gray, "baseline_multiplier": baseline_multiplier}


def _settings_highlight_rule(settings: dict) -> dict[str, Any]:
    tool_settings = settings.get(GUI_SETTINGS_KEY)
    if tool_settings is None:
        return _default_highlight_rule()
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    return _normalize_highlight_rule(tool_settings.get(HIGHLIGHT_RULE_KEY))


def _highlight_rule_from_state(state: GuiState, settings: dict) -> dict[str, Any]:
    return _normalize_highlight_rule(
        {
            "mode": state.highlight_mode.strip() or _settings_highlight_rule(settings)["mode"],
            "fixed_gray": state.highlight_fixed_gray.strip() or _settings_highlight_rule(settings)["fixed_gray"],
            "baseline_multiplier": state.highlight_baseline_multiplier.strip()
            or _settings_highlight_rule(settings)["baseline_multiplier"],
        }
    )


def _settings_excel_output_dir(settings: dict) -> Path:
    tool_settings = settings.get(GUI_SETTINGS_KEY)
    if tool_settings is None:
        return DEFAULT_EXCEL_OUTPUT_DIR
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    value = tool_settings.get(EXCEL_OUTPUT_DIR_KEY)
    return Path(str(value)) if value else DEFAULT_EXCEL_OUTPUT_DIR


def save_visual_analysis_settings(
    settings_path: Path,
    definitions: dict[str, dict[str, Any]],
    highlight_rule: dict[str, Any],
    excel_output_dir: Path,
) -> None:
    settings = _load_settings(settings_path)
    tool_settings = settings.setdefault(GUI_SETTINGS_KEY, {})
    if not isinstance(tool_settings, dict):
        raise ValueError(f"settings.{GUI_SETTINGS_KEY} must be an object")
    persisted_definitions: dict[str, dict[str, Any]] = {}
    persisted_rects: dict[str, dict[str, int]] = {}
    for roi_name in ROI_NAMES:
        definition = definitions.get(roi_name)
        if not definition:
            continue
        rect = tuple(definition["rect"])
        shape = _normalize_roi_shape(definition.get("shape", "rectangle"), roi_name)
        payload = _rect_to_settings_payload(rect)
        payload["shape"] = shape
        persisted_definitions[roi_name] = payload
        persisted_rects[roi_name] = _rect_to_settings_payload(rect)
    tool_settings[ROI_DEFINITIONS_KEY] = persisted_definitions
    tool_settings[ROI_RECT_OVERRIDES_KEY] = persisted_rects
    tool_settings[HIGHLIGHT_RULE_KEY] = _normalize_highlight_rule(highlight_rule)
    tool_settings[EXCEL_OUTPUT_DIR_KEY] = str(excel_output_dir)
    settings_path.write_text(json.dumps(settings, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _settings_roi_params(settings: dict, key: str, default: dict, label: str) -> dict:
    peak = settings.get("peak_detect")
    if not isinstance(peak, dict):
        return dict(default)
    roi = peak.get(key)
    if roi is None:
        return dict(default)
    if not isinstance(roi, dict):
        raise ValueError(f"settings.peak_detect.{key} must be an object")
    for param in ("left", "right", "top", "bottom"):
        if param not in roi:
            raise ValueError(f"settings.peak_detect.{key}.{param} is required")
    return {param: int(roi[param]) for param in ("left", "right", "top", "bottom")}


def _settings_roi2_params(settings: dict) -> dict:
    return _settings_roi_params(settings, "roi2_extension_params", ROI2_DEFAULT_PARAMS, "ROI2")


def _settings_roi3_params(settings: dict) -> dict:
    return _settings_roi_params(settings, "roi3_extension_params", ROI3_DEFAULT_PARAMS, "ROI3")


def _settings_roi4_rect(settings: dict) -> Optional[tuple[int, int, int, int]]:
    peak = settings.get("peak_detect")
    if not isinstance(peak, dict):
        return None
    return api_server.parse_roi4_rect(peak)


def _settings_roi4_bottom_region_ratio(settings: dict) -> Optional[float]:
    peak = settings.get("peak_detect")
    if not isinstance(peak, dict):
        return None
    roi4_rect = api_server.parse_roi4_rect(peak)
    selector = peak.get("roi4_after_selector")
    selector_enabled = bool(selector.get("enabled", False)) if isinstance(selector, dict) else False
    return api_server.parse_roi4_bottom_region_ratio(peak, selector_enabled, roi4_rect is not None)


def _settings_difference_threshold(settings: dict) -> float:
    peak = settings.get("peak_detect")
    if not isinstance(peak, dict):
        return 0.5
    threshold = peak.get("difference_threshold", 0.5)
    return float(threshold)


def _settings_focus_y_offset(settings: dict) -> float:
    focus_guides = settings.get("focus_guides")
    if not isinstance(focus_guides, dict):
        return 0.0
    return api_server.validate_focus_y_offset_mm(focus_guides.get("y_offset_mm", 0.0))


def _parse_roi2_params(text: Optional[str], settings: dict) -> dict:
    if not text:
        return _settings_roi2_params(settings)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError("--roi2-extension-params must be JSON, e.g. {\"left\":40,\"right\":40,\"top\":50,\"bottom\":30}") from exc
    if not isinstance(payload, dict):
        raise ValueError("--roi2-extension-params must be a JSON object")
    for key in ("left", "right", "top", "bottom"):
        if key not in payload:
            raise ValueError(f"--roi2-extension-params missing key: {key}")
    return payload


def _optional_path(text: str) -> Optional[Path]:
    stripped = text.strip()
    return Path(stripped) if stripped else None


def _optional_float(text: str, name: str) -> Optional[float]:
    stripped = text.strip()
    if not stripped:
        return None
    try:
        return float(stripped)
    except ValueError as exc:
        raise ValueError(f"{name} must be a number") from exc


def _optional_int(text: str, name: str) -> Optional[int]:
    stripped = text.strip()
    if not stripped:
        return None
    try:
        return int(stripped)
    except ValueError as exc:
        raise ValueError(f"{name} must be an integer") from exc


def _required_int(text: str, name: str) -> int:
    value = _optional_int(text, name)
    if value is None:
        raise ValueError(f"{name} is required")
    return value


def _roi_params_from_state(
    left: str,
    right: str,
    top: str,
    bottom: str,
    label: str,
    default_params: dict,
) -> dict:
    values = [left, right, top, bottom]
    if all(not str(value).strip() for value in values):
        return dict(default_params)
    return {
        "left": _required_int(left, f"{label}左扩展"),
        "right": _required_int(right, f"{label}右扩展"),
        "top": _required_int(top, f"{label}上扩展"),
        "bottom": _required_int(bottom, f"{label}下扩展"),
    }


def _roi4_rect_from_state(state: GuiState, settings: dict) -> Optional[tuple[int, int, int, int]]:
    fields = [state.roi4_x, state.roi4_y, state.roi4_width, state.roi4_height]
    if all(not str(value).strip() for value in fields):
        return _settings_roi4_rect(settings)
    if any(not str(value).strip() for value in fields):
        raise ValueError("ROI4固定区域需要同时填写X、Y、宽、高，或全部留空")
    x = _required_int(state.roi4_x, "ROI4 X")
    y = _required_int(state.roi4_y, "ROI4 Y")
    width = _required_int(state.roi4_width, "ROI4宽")
    height = _required_int(state.roi4_height, "ROI4高")
    if x < 0 or y < 0:
        raise ValueError("ROI4固定区域X/Y必须>=0")
    if width <= 0 or height <= 0:
        raise ValueError("ROI4固定区域宽/高必须>0")
    return x, y, x + width, y + height


def _roi4_bottom_region_ratio_from_state(state: GuiState, settings: dict, roi4_rect: Optional[tuple[int, int, int, int]]) -> Optional[float]:
    if state.roi4_bottom_region_ratio.strip():
        if roi4_rect is not None:
            raise ValueError("ROI4固定区域和底部高度比例不能同时填写")
        return api_server.validate_roi4_bottom_region_ratio(state.roi4_bottom_region_ratio)
    if roi4_rect is not None:
        return None
    return _settings_roi4_bottom_region_ratio(settings)


def config_from_gui_state(state: GuiState) -> AnalyzerConfig:
    root_text = state.root_dir.strip()
    output_text = state.output_csv.strip()
    focus_point_text = state.focus_point.strip()
    focus_points_csv = _optional_path(state.focus_points_csv)
    if not root_text:
        raise ValueError("序列根目录必填")
    if not output_text:
        raise ValueError("汇总CSV必填")
    if not focus_point_text and focus_points_csv is None:
        raise ValueError("必须填写全局焦点或选择焦点CSV")
    settings_path = _optional_path(state.settings_path)
    settings = _load_settings(settings_path) if settings_path is not None else {}
    roi2_params = _roi_params_from_state(
        state.roi2_left,
        state.roi2_right,
        state.roi2_top,
        state.roi2_bottom,
        "ROI2",
        _settings_roi2_params(settings),
    )
    roi3_params = _roi_params_from_state(
        state.roi3_left,
        state.roi3_right,
        state.roi3_top,
        state.roi3_bottom,
        "ROI3",
        _settings_roi3_params(settings),
    )
    roi4_rect = _roi4_rect_from_state(state, settings)
    roi4_bottom_region_ratio = _roi4_bottom_region_ratio_from_state(state, settings, roi4_rect)
    roi_definitions = _roi_definitions_from_inputs(state.roi_rect_inputs, state.roi_shape_inputs)
    roi_rect_overrides = _roi_rect_overrides_from_definitions(roi_definitions)
    highlight_rule = _highlight_rule_from_state(state, settings)
    threshold = _optional_float(state.difference_threshold, "差值阈值")
    focus_y_offset = _optional_float(state.focus_y_offset_mm, "焦点y偏移mm")
    provider_depth = _optional_float(state.provider_depth_mm, "超声深度mm")
    before_index = _required_int(state.before_frame_index, "基准帧序号")
    max_sequences = _optional_int(state.max_sequences, "最大序列数")
    after_strategy_text = state.after_strategy.strip() or "roi2_peak"
    after_strategy = AFTER_STRATEGY_LABELS.get(after_strategy_text, after_strategy_text)
    if after_strategy not in {"roi2_peak", "last"}:
        raise ValueError(f"不支持的治疗后帧选择：{after_strategy_text}")
    return AnalyzerConfig(
        root_dir=Path(root_text),
        output_csv=Path(output_text),
        per_frame_csv=_optional_path(state.per_frame_csv),
        settings_path=settings_path,
        focus_point=focus_point_text or None,
        focus_points_csv=focus_points_csv,
        provider_depth_mm=provider_depth,
        focus_y_offset_mm=api_server.validate_focus_y_offset_mm(
            focus_y_offset if focus_y_offset is not None else _settings_focus_y_offset(settings)
        ),
        roi2_extension_params=roi2_params,
        difference_threshold=float(threshold if threshold is not None else _settings_difference_threshold(settings)),
        before_frame_index=before_index,
        after_strategy=after_strategy,
        include_selected_debug=bool(state.include_selected_debug),
        max_sequences=max_sequences,
        roi3_extension_params=roi3_params,
        roi4_rect=roi4_rect,
        roi4_bottom_region_ratio=roi4_bottom_region_ratio,
        roi_rect_overrides=roi_rect_overrides,
        roi_definitions=roi_definitions,
        highlight_rule=highlight_rule,
        excel_output_dir=Path(state.excel_output_dir.strip()) if state.excel_output_dir.strip() else _settings_excel_output_dir(settings),
    )


def load_focus_points_csv(path: Optional[Path]) -> dict[str, Any]:
    if path is None:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"focus points CSV not found: {path}")
    result: dict[str, Any] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        required = {"sequence", "focus_point"}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"focus points CSV missing columns: {', '.join(sorted(missing))}")
        for row in reader:
            sequence = (row.get("sequence") or "").strip()
            focus_point = (row.get("focus_point") or "").strip()
            if sequence and focus_point:
                result[sequence] = focus_point
    return result


def list_sequences(root_dir: Path, max_sequences: Optional[int]) -> list[Path]:
    if not root_dir.exists():
        raise FileNotFoundError(f"root directory not found: {root_dir}")
    if not root_dir.is_dir():
        raise NotADirectoryError(f"root path is not a directory: {root_dir}")
    sequences = sorted([p for p in root_dir.iterdir() if p.is_dir()], key=lambda p: p.name)
    if max_sequences is not None:
        return sequences[: int(max_sequences)]
    return sequences


def list_frame_paths(sequence_dir: Path, include_selected_debug: bool) -> list[Path]:
    frames = []
    raw_frame_pattern = re.compile(r"^\d+_.*_frame\.png$", re.IGNORECASE)
    for path in sequence_dir.iterdir():
        if not path.is_file() or path.suffix.lower() != ".png":
            continue
        is_raw_frame = bool(raw_frame_pattern.match(path.name))
        is_selected_debug = path.name.lower().startswith("selected_")
        if not is_raw_frame and not (include_selected_debug and is_selected_debug):
            continue
        frames.append(path)
    return sorted(frames, key=lambda p: p.name)


def load_frame(path: Path) -> np.ndarray:
    with Image.open(path) as im:
        return np.asarray(im.convert("RGB"), dtype=np.uint8)


def frame_roi2_mean(path: Path, roi2_rect: tuple[int, int, int, int]) -> float:
    frame = load_frame(path)
    return api_server.roi_gray_mean(frame, roi2_rect)


def _empty_roi_stats() -> dict[str, Any]:
    return {
        "rect": "",
        "shape": "",
        "width": "",
        "height": "",
        "area": "",
        "mean": "",
        "density": "",
        "std": "",
        "median": "",
        "median_abs_deviation": "",
        "skewness": "",
        "kurtosis": "",
        "p01": "",
        "p10": "",
        "p50": "",
        "p90": "",
        "p99": "",
        "threshold": "",
        "highlight_count": "",
        "highlight_area": "",
        "highlight_ratio": "",
        "highlight_std": "",
        "hem_z_count": "",
        "hem_z_area": "",
        "mean_delta": "",
        "mean_delta_pct": "",
        "std_delta": "",
        "median_delta": "",
        "highlight_area_delta": "",
        "histogram": [0] * 256,
    }


def _roi_gray_pixels(frame: np.ndarray, rect: tuple[int, int, int, int], shape: str = "rectangle") -> np.ndarray:
    x1, y1, x2, y2 = [int(v) for v in rect]
    roi = frame[y1:y2, x1:x2]
    if roi.size == 0:
        return np.asarray([], dtype=np.float64)
    gray = api_server.gray_image(roi).astype(np.float64)
    shape = _normalize_roi_shape(shape, "ROI")
    if shape == "rectangle":
        return gray.reshape(-1)
    height, width = gray.shape[:2]
    if width <= 0 or height <= 0:
        return np.asarray([], dtype=np.float64)
    yy, xx = np.ogrid[:height, :width]
    cx = (width - 1) / 2.0
    cy = (height - 1) / 2.0
    rx = max(width / 2.0, 0.5)
    ry = max(height / 2.0, 0.5)
    mask = (((xx - cx) / rx) ** 2 + ((yy - cy) / ry) ** 2) <= 1.0
    return gray[mask].reshape(-1)


def roi_gray_histogram(
    frame: np.ndarray,
    rect: Optional[tuple[int, int, int, int]],
    shape: str = "rectangle",
) -> list[int]:
    if rect is None:
        return [0] * 256
    pixels = _roi_gray_pixels(frame, rect, shape)
    if pixels.size == 0:
        return [0] * 256
    bins = np.clip(np.rint(pixels), 0, 255).astype(np.uint8)
    return np.bincount(bins, minlength=256).astype(int).tolist()


def _highlight_threshold(
    mean: float,
    baseline_mean: Optional[float],
    highlight_rule: Optional[dict[str, Any]],
) -> float:
    rule = _normalize_highlight_rule(highlight_rule)
    if rule["mode"] == "fixed_gray":
        return float(rule["fixed_gray"])
    threshold_base = baseline_mean if baseline_mean is not None else mean
    return float(threshold_base) * float(rule["baseline_multiplier"])


def roi_stats_for_frame(
    frame: np.ndarray,
    rect: Optional[tuple[int, int, int, int]],
    baseline_stats: Optional[dict[str, str]] = None,
    shape: str = "rectangle",
    highlight_rule: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    if rect is None:
        return _empty_roi_stats()
    shape = _normalize_roi_shape(shape, "ROI")
    x1, y1, x2, y2 = [int(v) for v in rect]
    width = max(0, x2 - x1)
    height = max(0, y2 - y1)
    pixels = _roi_gray_pixels(frame, (x1, y1, x2, y2), shape)
    area = int(pixels.size)
    if pixels.size == 0:
        stats = _empty_roi_stats()
        stats.update({
            "rect": _fmt_rect((x1, y1, x2, y2)),
            "shape": shape,
            "width": str(width),
            "height": str(height),
            "area": "0",
        })
        return stats
    histogram = roi_gray_histogram(frame, (x1, y1, x2, y2), shape)
    mean = float(np.mean(pixels))
    std = float(np.std(pixels))
    median = float(np.median(pixels))
    median_abs_deviation = float(np.median(np.abs(pixels - median)))
    centered = pixels - mean
    if std > 0:
        skewness = float(np.mean((centered / std) ** 3))
        kurtosis = float(np.mean((centered / std) ** 4))
    else:
        skewness = 0.0
        kurtosis = 0.0
    p01, p10, p50, p90, p99 = [float(v) for v in np.percentile(pixels, [1, 10, 50, 90, 99])]
    baseline_mean = _optional_float((baseline_stats or {}).get("mean", ""), "baseline mean") if baseline_stats else None
    baseline_std = _optional_float((baseline_stats or {}).get("std", ""), "baseline std") if baseline_stats else None
    threshold = _highlight_threshold(mean, baseline_mean, highlight_rule)
    highlight_pixels = pixels[pixels >= threshold]
    highlight_count = int(highlight_pixels.size)
    highlight_ratio = float(highlight_count / area) if area > 0 else 0.0
    highlight_std = float(np.std(highlight_pixels)) if highlight_count else 0.0
    hem_z_count = 0
    if baseline_mean is not None and baseline_std is not None and baseline_std > 0:
        hem_z_count = int(np.count_nonzero(((pixels - baseline_mean) / baseline_std) >= HEM_Z_SCORE_THRESHOLD))
    stats = {
        "rect": _fmt_rect((x1, y1, x2, y2)),
        "shape": shape,
        "width": str(width),
        "height": str(height),
        "area": str(area),
        "mean": _fmt_float(mean),
        "density": _fmt_float(mean / 255.0),
        "std": _fmt_float(std),
        "median": _fmt_float(median),
        "median_abs_deviation": _fmt_float(median_abs_deviation),
        "skewness": _fmt_float(skewness),
        "kurtosis": _fmt_float(kurtosis),
        "p01": _fmt_float(p01),
        "p10": _fmt_float(p10),
        "p50": _fmt_float(p50),
        "p90": _fmt_float(p90),
        "p99": _fmt_float(p99),
        "threshold": _fmt_float(threshold),
        "highlight_count": str(highlight_count),
        "highlight_area": str(highlight_count),
        "highlight_ratio": _fmt_float(highlight_ratio),
        "highlight_std": _fmt_float(highlight_std),
        "hem_z_count": str(hem_z_count),
        "hem_z_area": str(hem_z_count),
        "mean_delta": "",
        "mean_delta_pct": "",
        "std_delta": "",
        "median_delta": "",
        "highlight_area_delta": "",
        "histogram": histogram,
    }
    if baseline_stats:
        baseline_mean = _optional_float(baseline_stats.get("mean", ""), "baseline mean")
        baseline_std_value = _optional_float(baseline_stats.get("std", ""), "baseline std")
        baseline_median = _optional_float(baseline_stats.get("median", ""), "baseline median")
        baseline_highlight_area = _optional_int(baseline_stats.get("highlight_area", ""), "baseline highlight area")
        if baseline_mean is not None:
            stats["mean_delta"] = _fmt_float(mean - baseline_mean)
            if baseline_mean != 0:
                stats["mean_delta_pct"] = _fmt_float((mean - baseline_mean) / baseline_mean)
        if baseline_std_value is not None:
            stats["std_delta"] = _fmt_float(std - baseline_std_value)
        if baseline_median is not None:
            stats["median_delta"] = _fmt_float(median - baseline_median)
        if baseline_highlight_area is not None:
            stats["highlight_area_delta"] = str(highlight_count - baseline_highlight_area)
    return stats


def roi_stats_for_frame_set(
    frame: np.ndarray,
    roi_meta: dict[str, Any],
    baseline_stats: Optional[dict[str, dict[str, str]]] = None,
) -> dict[str, dict[str, Any]]:
    shapes = roi_meta.get("roi_shapes", {})
    highlight_rule = roi_meta.get("highlight_rule")
    return {
        "ROI1": roi_stats_for_frame(frame, roi_meta["roi1_rect"], (baseline_stats or {}).get("ROI1"), shapes.get("ROI1", "rectangle"), highlight_rule),
        "ROI2": roi_stats_for_frame(frame, roi_meta["roi2_rect"], (baseline_stats or {}).get("ROI2"), shapes.get("ROI2", "rectangle"), highlight_rule),
        "ROI3": roi_stats_for_frame(frame, roi_meta["roi3_rect"], (baseline_stats or {}).get("ROI3"), shapes.get("ROI3", "rectangle"), highlight_rule),
        "ROI4": roi_stats_for_frame(frame, roi_meta["roi4_rect"], (baseline_stats or {}).get("ROI4"), shapes.get("ROI4", "rectangle"), highlight_rule),
    }


def _scale_rect(rect: tuple[int, int, int, int], scale: float) -> tuple[int, int, int, int]:
    return tuple(int(round(v * scale)) for v in rect)


def _scale_point(point: tuple[int, int], scale: float) -> tuple[int, int]:
    return int(round(point[0] * scale)), int(round(point[1] * scale))


def _draw_focus_cross(draw: ImageDraw.ImageDraw, point: tuple[int, int], radius: int = 7) -> None:
    x, y = point
    color = FOCUS_COLOR
    draw.ellipse((x - radius, y - radius, x + radius, y + radius), outline=color, width=3)
    draw.line((x - radius - 4, y, x + radius + 4, y), fill=color, width=2)
    draw.line((x, y - radius - 4, x, y + radius + 4), fill=color, width=2)


def _draw_scaled_rect(
    draw: ImageDraw.ImageDraw,
    rect: tuple[int, int, int, int],
    scale: float,
    color: tuple[int, int, int],
    width: int,
    label: str,
) -> None:
    x1, y1, x2, y2 = _scale_rect(rect, scale)
    x2 = max(x1 + 1, x2 - 1)
    y2 = max(y1 + 1, y2 - 1)
    draw.rectangle((x1, y1, x2, y2), outline=color, width=width)
    draw.text((x1 + 4, y1 + 4), label, fill=color)


def _draw_scaled_roi_shape(
    draw: ImageDraw.ImageDraw,
    rect: tuple[int, int, int, int],
    shape: str,
    scale: float,
    color: tuple[int, int, int],
    width: int,
    label: str,
) -> None:
    x1, y1, x2, y2 = _scale_rect(rect, scale)
    x2 = max(x1 + 1, x2 - 1)
    y2 = max(y1 + 1, y2 - 1)
    if _normalize_roi_shape(shape, label) == "ellipse":
        draw.ellipse((x1, y1, x2, y2), outline=color, width=width)
    else:
        draw.rectangle((x1, y1, x2, y2), outline=color, width=width)
    draw.text((x1 + 4, y1 + 4), label, fill=color)


def render_roi_histogram_image(histogram: list[int], size: tuple[int, int] = (260, 48)) -> Image.Image:
    width, height = size
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width - 1, height - 1), outline=(210, 210, 210), width=1)
    values = [int(v or 0) for v in histogram[:256]]
    if len(values) < 256:
        values.extend([0] * (256 - len(values)))
    max_count = max(values) if values else 0
    if max_count <= 0:
        return image
    plot_height = max(1, height - 4)
    for bin_index, count in enumerate(values):
        x = int(round(bin_index * (width - 1) / 255))
        bar_height = int(round((count / max_count) * plot_height))
        draw.line((x, height - 2, x, height - 2 - bar_height), fill=(80, 120, 180))
    return image


def render_sequence_preview_image(
    frame_path: Path,
    sequence_name: str,
    config: AnalyzerConfig,
    focus_points: dict[str, Any],
    show_roi2: bool,
    show_focus: bool,
    max_size: tuple[int, int] = PREVIEW_MAX_SIZE,
    show_roi1: bool = False,
    show_roi3: bool = False,
    show_roi4: bool = False,
    baseline_frame_path: Optional[Path] = None,
) -> tuple[Image.Image, dict[str, Any]]:
    frame = load_frame(frame_path)
    focus_anchor = resolve_sequence_focus(sequence_name, config, focus_points)
    roi_meta = resolve_roi_rects(
        frame,
        focus_anchor,
        config,
        include_roi3=show_roi3,
        include_roi4=show_roi4,
    )
    baseline_stats = None
    if baseline_frame_path is not None:
        baseline_frame = load_frame(baseline_frame_path)
        baseline_stats = roi_stats_for_frame_set(baseline_frame, roi_meta)
    image = Image.fromarray(frame).convert("RGB")
    original_width, original_height = image.size
    max_width = max(1, int(max_size[0]))
    max_height = max(1, int(max_size[1]))
    scale = min(max_width / original_width, max_height / original_height)
    if scale != 1.0:
        image = image.resize((int(round(original_width * scale)), int(round(original_height * scale))), Image.Resampling.LANCZOS)
    draw = ImageDraw.Draw(image)
    overlay_width = max(2, int(round(3 * scale)))
    shapes = roi_meta.get("roi_shapes", {})
    if show_roi1:
        _draw_scaled_roi_shape(draw, roi_meta["roi1_rect"], shapes.get("ROI1", "rectangle"), scale, ROI1_COLOR, overlay_width, "ROI1")
    if show_roi4 and roi_meta["roi4_rect"] is not None:
        _draw_scaled_roi_shape(draw, roi_meta["roi4_rect"], shapes.get("ROI4", "rectangle"), scale, ROI4_COLOR, overlay_width, "ROI4")
    if show_roi3:
        _draw_scaled_roi_shape(draw, roi_meta["roi3_rect"], shapes.get("ROI3", "rectangle"), scale, ROI3_COLOR, overlay_width, "ROI3")
    if show_roi2:
        _draw_scaled_roi_shape(draw, roi_meta["roi2_rect"], shapes.get("ROI2", "rectangle"), scale, ROI2_COLOR, overlay_width, "ROI2")
    if show_focus:
        _draw_focus_cross(draw, _scale_point(focus_anchor, scale), radius=max(7, int(round(7 * scale))))
    meta = {
        "focus_anchor": focus_anchor,
        "scale": scale,
        "frame_path": frame_path,
        "roi_stats": roi_stats_for_frame_set(frame, roi_meta, baseline_stats),
    }
    meta.update(roi_meta)
    return image, meta


def resolve_sequence_focus(sequence_name: str, config: AnalyzerConfig, focus_points: dict[str, Any]) -> tuple[int, int]:
    raw_focus = focus_points.get(sequence_name, config.focus_point)
    anchor = _parse_focus_point_value(raw_focus)
    if anchor is None:
        raise ValueError(f"focus_point is required for sequence {sequence_name}")
    return anchor


def resolve_roi2_rect(
    first_frame: np.ndarray,
    anchor: tuple[int, int],
    config: AnalyzerConfig,
) -> tuple[tuple[int, int], tuple[int, int, int, int]]:
    height, width = first_frame.shape[:2]
    offline_config = api_server.OfflineConfig(
        roi2_extension_params=dict(config.roi2_extension_params),
        focus_y_offset_mm=float(config.focus_y_offset_mm),
    )
    offset_anchor = api_server.resolve_offset_focus_anchor(
        (width, height),
        anchor,
        config.provider_depth_mm,
        offline_config,
    )
    override = config.roi_rect_overrides.get("ROI2")
    if override is not None:
        return offset_anchor, _validate_roi_rect_for_image(override, width, height, "ROI2")
    roi2_rect = api_server.compute_roi_region(
        (width, height),
        offset_anchor,
        config.roi2_extension_params,
    )
    if roi2_rect is None:
        raise ValueError(
            "ROI2 rectangle is outside image bounds "
            f"sequence_size={(width, height)} focus={anchor} offset_anchor={offset_anchor} params={config.roi2_extension_params}"
        )
    return offset_anchor, roi2_rect


def resolve_roi_rects(
    first_frame: np.ndarray,
    anchor: tuple[int, int],
    config: AnalyzerConfig,
    include_roi3: bool = True,
    include_roi4: bool = True,
) -> dict[str, Any]:
    height, width = first_frame.shape[:2]
    offline_config = api_server.OfflineConfig(
        roi2_extension_params=dict(config.roi2_extension_params),
        roi3_extension_params=dict(config.roi3_extension_params),
        roi4_rect=config.roi4_rect,
        roi4_bottom_region_ratio=config.roi4_bottom_region_ratio,
        focus_y_offset_mm=float(config.focus_y_offset_mm),
    )
    offset_anchor = api_server.resolve_offset_focus_anchor(
        (width, height),
        anchor,
        config.provider_depth_mm,
        offline_config,
    )
    roi2_rect = config.roi_rect_overrides.get("ROI2")
    if roi2_rect is not None:
        roi2_rect = _validate_roi_rect_for_image(roi2_rect, width, height, "ROI2")
    else:
        roi2_rect = api_server.compute_roi_region(
            (width, height),
            offset_anchor,
            config.roi2_extension_params,
        )
    if roi2_rect is None:
        raise ValueError(
            "ROI2 rectangle is outside image bounds "
            f"sequence_size={(width, height)} focus={anchor} offset_anchor={offset_anchor} params={config.roi2_extension_params}"
        )
    roi3_rect = None
    if include_roi3:
        roi3_rect = config.roi_rect_overrides.get("ROI3")
        if roi3_rect is not None:
            roi3_rect = _validate_roi_rect_for_image(roi3_rect, width, height, "ROI3")
        else:
            roi3_rect = api_server.compute_roi_region(
                (width, height),
                offset_anchor,
                config.roi3_extension_params,
            )
        if roi3_rect is None:
            raise ValueError(
                "ROI3 rectangle is outside image bounds "
                f"sequence_size={(width, height)} focus={anchor} offset_anchor={offset_anchor} params={config.roi3_extension_params}"
            )
    roi4_rect = None
    if include_roi4:
        roi4_rect = config.roi_rect_overrides.get("ROI4")
        if roi4_rect is not None:
            roi4_rect = _validate_roi_rect_for_image(roi4_rect, width, height, "ROI4")
        else:
            roi4_rect = api_server.resolve_roi4_rect_for_image(offline_config, first_frame)
    roi1_rect = config.roi_rect_overrides.get("ROI1")
    if roi1_rect is not None:
        roi1_rect = _validate_roi_rect_for_image(roi1_rect, width, height, "ROI1")
    else:
        roi1_rect = (0, 0, width, height)
    roi_shapes = {
        roi_name: _normalize_roi_shape(
            (config.roi_definitions.get(roi_name) or {}).get("shape", "rectangle"),
            roi_name,
        )
        for roi_name in ROI_NAMES
    }
    return {
        "offset_anchor": offset_anchor,
        "roi1_rect": roi1_rect,
        "roi2_rect": roi2_rect,
        "roi3_rect": roi3_rect,
        "roi4_rect": roi4_rect,
        "roi_shapes": roi_shapes,
        "highlight_rule": _normalize_highlight_rule(config.highlight_rule),
    }


def choose_after_index(frame_means: list[float], before_index_zero_based: int, strategy: str) -> int:
    if not frame_means:
        raise ValueError("frame_means is empty")
    if strategy == "last":
        return len(frame_means) - 1
    if strategy == "roi2_peak":
        candidates = range(before_index_zero_based + 1, len(frame_means))
        best_index = None
        best_mean = None
        for index in candidates:
            mean = frame_means[index]
            if best_mean is None or mean > best_mean:
                best_index = index
                best_mean = mean
        if best_index is None:
            raise ValueError("after_strategy=roi2_peak requires at least one frame after before_frame_index")
        return best_index
    raise ValueError(f"unsupported after_strategy: {strategy}")


def analyze_sequence(
    sequence_dir: Path,
    config: AnalyzerConfig,
    focus_points: dict[str, Any],
) -> tuple[dict[str, str], list[dict[str, str]]]:
    frame_paths = list_frame_paths(sequence_dir, config.include_selected_debug)
    if not frame_paths:
        raise ValueError(f"sequence has no PNG frames: {sequence_dir}")
    before_index = int(config.before_frame_index) - 1
    if before_index < 0 or before_index >= len(frame_paths):
        raise ValueError(
            f"before_frame_index out of range for sequence {sequence_dir.name}: "
            f"{config.before_frame_index} not in 1..{len(frame_paths)}"
        )
    focus_anchor = resolve_sequence_focus(sequence_dir.name, config, focus_points)
    first_frame = load_frame(frame_paths[before_index])
    offset_anchor, roi2_rect = resolve_roi2_rect(first_frame, focus_anchor, config)

    frame_means = [frame_roi2_mean(path, roi2_rect) for path in frame_paths]
    before_mean = frame_means[before_index]
    after_index = choose_after_index(frame_means, before_index, config.after_strategy)
    after_mean = frame_means[after_index]
    roi2_diff = float(after_mean) - float(before_mean)
    roi2_color = "green" if roi2_diff >= float(config.difference_threshold) else "red"

    row = {
        "sequence": sequence_dir.name,
        "status": "ok",
        "error": "",
        "frame_count": str(len(frame_paths)),
        "focus_anchor": _fmt_point(focus_anchor),
        "offset_anchor": _fmt_point(offset_anchor),
        "roi2_rect": _fmt_rect(roi2_rect),
        "before_frame_index": str(before_index + 1),
        "before_frame": frame_paths[before_index].name,
        "before_mean": _fmt_float(before_mean),
        "after_frame_index": str(after_index + 1),
        "after_frame": frame_paths[after_index].name,
        "after_mean": _fmt_float(after_mean),
        "roi2_diff": _fmt_float(roi2_diff),
        "difference_threshold": _fmt_float(config.difference_threshold),
        "roi2_color": roi2_color,
        "after_strategy": config.after_strategy,
    }
    frame_rows = []
    for index, path in enumerate(frame_paths):
        frame_rows.append(
            {
                "sequence": sequence_dir.name,
                "frame_index": str(index + 1),
                "frame": path.name,
                "roi2_mean": _fmt_float(frame_means[index]),
                "roi2_diff_from_before": _fmt_float(float(frame_means[index]) - float(before_mean)),
            }
        )
    return row, frame_rows


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def analyze_root(
    config: AnalyzerConfig,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
) -> list[dict[str, str]]:
    focus_points = load_focus_points_csv(config.focus_points_csv)
    summary_rows: list[dict[str, str]] = []
    frame_rows: list[dict[str, str]] = []
    sequence_dirs = list_sequences(config.root_dir, config.max_sequences)
    total = len(sequence_dirs)
    for index, sequence_dir in enumerate(sequence_dirs, start=1):
        if progress_callback is not None:
            progress_callback(index, total, sequence_dir.name)
        row, sequence_frame_rows = analyze_sequence(sequence_dir, config, focus_points)
        summary_rows.append(row)
        frame_rows.extend(sequence_frame_rows)
    write_csv(config.output_csv, SUMMARY_FIELDS, summary_rows)
    if config.per_frame_csv is not None:
        write_csv(config.per_frame_csv, FRAME_FIELDS, frame_rows)
    return summary_rows


def _excel_stats_headers() -> list[str]:
    fields = [
        "sequence",
        "frame_index",
        "frame",
        "roi",
        "shape",
        "rect",
        "width",
        "height",
        "area",
        "mean",
        "density",
        "std",
        "median",
        "median_abs_deviation",
        "p10",
        "p90",
        "threshold",
        "highlight_count",
        "highlight_area",
        "highlight_ratio",
        "highlight_std",
        "hem_z_area",
        "mean_delta",
        "mean_delta_pct",
        "std_delta",
        "median_delta",
        "highlight_area_delta",
    ]
    return fields


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return ",".join(str(v) for v in value)
    if value == "":
        return ""
    if isinstance(value, str):
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            return value
    return value


def export_sequence_excel(
    sequence_dir: Path,
    config: AnalyzerConfig,
    focus_points: dict[str, Any],
    output_path: Path,
) -> Path:
    from openpyxl import Workbook

    frame_paths = list_frame_paths(sequence_dir, config.include_selected_debug)
    if not frame_paths:
        raise ValueError(f"sequence has no PNG frames: {sequence_dir}")
    before_index = int(config.before_frame_index) - 1
    if before_index < 0 or before_index >= len(frame_paths):
        raise ValueError(
            f"before_frame_index out of range for sequence {sequence_dir.name}: "
            f"{config.before_frame_index} not in 1..{len(frame_paths)}"
        )

    focus_anchor = resolve_sequence_focus(sequence_dir.name, config, focus_points)
    baseline_frame = load_frame(frame_paths[before_index])
    roi_meta = resolve_roi_rects(baseline_frame, focus_anchor, config, include_roi3=True, include_roi4=True)
    baseline_stats = roi_stats_for_frame_set(baseline_frame, roi_meta)
    summary_row, _frame_rows = analyze_sequence(sequence_dir, config, focus_points)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=False)
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["字段", "值"])
    summary_sheet.append(["sequence", sequence_dir.name])
    summary_sheet.append(["frame_count", len(frame_paths)])
    summary_sheet.append(["before_frame_index", config.before_frame_index])
    summary_sheet.append(["before_frame", frame_paths[before_index].name])
    summary_sheet.append(["highlight_mode", _normalize_highlight_rule(config.highlight_rule)["mode"]])
    summary_sheet.append(["fixed_gray", _normalize_highlight_rule(config.highlight_rule)["fixed_gray"]])
    summary_sheet.append(["baseline_multiplier", _normalize_highlight_rule(config.highlight_rule)["baseline_multiplier"]])
    summary_sheet.append(["roi2_color", summary_row["roi2_color"]])
    summary_sheet.append(["roi2_diff", summary_row["roi2_diff"]])
    summary_sheet.append(["summary_csv", str(config.output_csv)])

    stats_sheet = workbook.create_sheet("Frame_ROI_Stats")
    stats_headers = _excel_stats_headers()
    stats_sheet.append(stats_headers)

    histogram_sheet = workbook.create_sheet("Histograms")
    histogram_headers = ["sequence", "frame_index", "frame", "roi"] + [f"gray_{index}" for index in range(256)]
    histogram_sheet.append(histogram_headers)

    config_sheet = workbook.create_sheet("ROI_Config")
    config_sheet.append(["roi", "shape", "x", "y", "width", "height", "highlight_mode", "fixed_gray", "baseline_multiplier"])
    highlight_rule = _normalize_highlight_rule(config.highlight_rule)
    for roi_name in ROI_NAMES:
        rect = roi_meta.get(f"{roi_name.lower()}_rect")
        if rect is None:
            continue
        x1, y1, x2, y2 = [int(v) for v in rect]
        config_sheet.append(
            [
                roi_name,
                roi_meta["roi_shapes"].get(roi_name, "rectangle"),
                x1,
                y1,
                x2 - x1,
                y2 - y1,
                highlight_rule["mode"],
                highlight_rule["fixed_gray"],
                highlight_rule["baseline_multiplier"],
            ]
        )

    for frame_index, frame_path in enumerate(frame_paths, start=1):
        frame = load_frame(frame_path)
        stats_by_roi = roi_stats_for_frame_set(frame, roi_meta, baseline_stats)
        for roi_name in ROI_NAMES:
            stats = stats_by_roi[roi_name]
            stats_row = {
                "sequence": sequence_dir.name,
                "frame_index": frame_index,
                "frame": frame_path.name,
                "roi": roi_name,
                **stats,
            }
            stats_sheet.append([_excel_cell_value(stats_row.get(header, "")) for header in stats_headers])
            histogram_sheet.append(
                [sequence_dir.name, frame_index, frame_path.name, roi_name]
                + [int(value) for value in stats.get("histogram", [0] * 256)]
            )

    workbook.save(output_path)
    return output_path


def build_config_from_args(argv: Optional[list[str]] = None) -> AnalyzerConfig:
    parser = argparse.ArgumentParser(description="Batch analyze HEM ROI2 gray-difference metrics for treatment sequences.")
    parser.add_argument("--root", required=True, help="Root directory containing one treatment sequence per child folder.")
    parser.add_argument("--output-csv", required=True, help="Summary CSV output path.")
    parser.add_argument("--per-frame-csv", help="Optional per-frame ROI2 metric CSV output path.")
    parser.add_argument("--settings", default="settings", help="Settings JSON path used for default ROI2 params and thresholds.")
    parser.add_argument("--focus-point", help='Global focus point, e.g. "PointF(434.85052, 272.8398)" or "434,272".')
    parser.add_argument("--focus-points-csv", help="Optional CSV with columns: sequence,focus_point.")
    parser.add_argument("--provider-depth-mm", type=float, help="Provider ultrasound depth in mm, required when focus_y_offset_mm > 0.")
    parser.add_argument("--focus-y-offset-mm", type=float, help="Override focus_guides.y_offset_mm from settings.")
    parser.add_argument("--roi2-extension-params", help='Override ROI2 params as JSON, e.g. {"left":40,"right":40,"top":50,"bottom":30}.')
    parser.add_argument("--difference-threshold", type=float, help="Override ROI2 green threshold.")
    parser.add_argument("--before-frame-index", type=int, default=1, help="1-based before/baseline frame index. Default: 1.")
    parser.add_argument("--after-strategy", choices=["roi2_peak", "last"], default="roi2_peak", help="After frame selection strategy.")
    parser.add_argument("--include-selected-debug", action="store_true", help="Include selected_*.png debug images in frame analysis.")
    parser.add_argument("--max-sequences", type=int, help="Limit analyzed sequence count for smoke runs.")
    parser.add_argument("--gui", action="store_true", help="Open the single-file GUI instead of running CLI analysis.")
    args = parser.parse_args(argv)

    settings_path = Path(args.settings) if args.settings else None
    settings = _load_settings(settings_path) if settings_path is not None else {}
    focus_y_offset = (
        api_server.validate_focus_y_offset_mm(args.focus_y_offset_mm)
        if args.focus_y_offset_mm is not None
        else _settings_focus_y_offset(settings)
    )
    roi2_params = _parse_roi2_params(args.roi2_extension_params, settings)
    roi4_rect = _settings_roi4_rect(settings)
    roi_definitions = _settings_roi_definitions(settings)
    difference_threshold = (
        float(args.difference_threshold)
        if args.difference_threshold is not None
        else _settings_difference_threshold(settings)
    )
    return AnalyzerConfig(
        root_dir=Path(args.root),
        output_csv=Path(args.output_csv),
        per_frame_csv=Path(args.per_frame_csv) if args.per_frame_csv else None,
        settings_path=settings_path,
        focus_point=args.focus_point,
        focus_points_csv=Path(args.focus_points_csv) if args.focus_points_csv else None,
        provider_depth_mm=args.provider_depth_mm,
        focus_y_offset_mm=focus_y_offset,
        roi2_extension_params=roi2_params,
        difference_threshold=difference_threshold,
        before_frame_index=int(args.before_frame_index),
        after_strategy=args.after_strategy,
        include_selected_debug=bool(args.include_selected_debug),
        max_sequences=args.max_sequences,
        roi3_extension_params=_settings_roi3_params(settings),
        roi4_rect=roi4_rect,
        roi4_bottom_region_ratio=_settings_roi4_bottom_region_ratio(settings) if roi4_rect is None else None,
        roi_rect_overrides=_roi_rect_overrides_from_definitions(roi_definitions),
        roi_definitions=roi_definitions,
        highlight_rule=_settings_highlight_rule(settings),
        excel_output_dir=_settings_excel_output_dir(settings),
    )


class HemRoi2BatchAnalyzerGui:
    def __init__(self, root):
        import tkinter as tk
        from tkinter import ttk

        self.root = root
        self.tk = tk
        self.ttk = ttk
        self.root.title("HEM ROI2 单序列可视化分析器")
        self._maximize_root()
        self.root_dir = tk.StringVar(value="E:\\20260614")
        self.output_csv = tk.StringVar(value=str(Path("doc") / "tasks" / "hem-roi2-batch-analyzer" / "summary.csv"))
        self.per_frame_csv = tk.StringVar(value=str(Path("doc") / "tasks" / "hem-roi2-batch-analyzer" / "frames.csv"))
        self.settings_path = tk.StringVar(value="settings")
        self.focus_point = tk.StringVar(value=DEFAULT_FOCUS_POINT)
        focus_x, focus_y = _parse_focus_point_value(DEFAULT_FOCUS_POINT)
        self.focus_x = tk.StringVar(value=str(focus_x))
        self.focus_y = tk.StringVar(value=str(focus_y))
        self.focus_points_csv = tk.StringVar(value="")
        self.provider_depth_mm = tk.StringVar(value="")
        self.focus_y_offset_mm = tk.StringVar(value=str(_settings_focus_y_offset(_load_settings(Path("settings"))) if Path("settings").exists() else 0.0))
        settings = _load_settings(Path("settings")) if Path("settings").exists() else {}
        roi2 = _settings_roi2_params(settings)
        roi3 = _settings_roi3_params(settings)
        roi4_rect = _settings_roi4_rect(settings)
        roi4_ratio = _settings_roi4_bottom_region_ratio(settings) if roi4_rect is None else None
        roi_definitions = _settings_roi_definitions(settings)
        roi_rect_overrides = _roi_rect_overrides_from_definitions(roi_definitions)
        highlight_rule = _settings_highlight_rule(settings)
        self.roi2_left = tk.StringVar(value=str(roi2["left"]))
        self.roi2_right = tk.StringVar(value=str(roi2["right"]))
        self.roi2_top = tk.StringVar(value=str(roi2["top"]))
        self.roi2_bottom = tk.StringVar(value=str(roi2["bottom"]))
        self.roi3_left = tk.StringVar(value=str(roi3["left"]))
        self.roi3_right = tk.StringVar(value=str(roi3["right"]))
        self.roi3_top = tk.StringVar(value=str(roi3["top"]))
        self.roi3_bottom = tk.StringVar(value=str(roi3["bottom"]))
        if roi4_rect is None:
            roi4_x = roi4_y = roi4_width = roi4_height = ""
        else:
            roi4_x = str(roi4_rect[0])
            roi4_y = str(roi4_rect[1])
            roi4_width = str(roi4_rect[2] - roi4_rect[0])
            roi4_height = str(roi4_rect[3] - roi4_rect[1])
        self.roi4_x = tk.StringVar(value=roi4_x)
        self.roi4_y = tk.StringVar(value=roi4_y)
        self.roi4_width = tk.StringVar(value=roi4_width)
        self.roi4_height = tk.StringVar(value=roi4_height)
        self.roi4_bottom_region_ratio = tk.StringVar(value="" if roi4_ratio is None else str(roi4_ratio))
        self.roi_rect_vars = {
            roi_name: {
                field_name: tk.StringVar(value=value)
                for field_name, value in _rect_to_input_values(roi_rect_overrides.get(roi_name)).items()
            }
            for roi_name in ROI_NAMES
        }
        self.roi_shape_vars = {
            roi_name: tk.StringVar(value=ROI_SHAPE_DISPLAY[(roi_definitions.get(roi_name) or {}).get("shape", "rectangle")])
            for roi_name in ROI_NAMES
        }
        self.highlight_mode = tk.StringVar(value=HIGHLIGHT_MODE_DISPLAY[highlight_rule["mode"]])
        self.highlight_fixed_gray = tk.StringVar(value=str(highlight_rule["fixed_gray"]))
        self.highlight_baseline_multiplier = tk.StringVar(value=str(highlight_rule["baseline_multiplier"]))
        self.excel_output_dir = tk.StringVar(value=str(_settings_excel_output_dir(settings)))
        self.difference_threshold = tk.StringVar(value=str(_settings_difference_threshold(settings)))
        self.before_frame_index = tk.StringVar(value="1")
        self.after_strategy = tk.StringVar(value=AFTER_STRATEGY_VALUES["roi2_peak"])
        self.include_selected_debug = tk.BooleanVar(value=False)
        self.show_roi1 = tk.BooleanVar(value=True)
        self.show_roi2 = tk.BooleanVar(value=True)
        self.show_roi3 = tk.BooleanVar(value=True)
        self.show_roi4 = tk.BooleanVar(value=True)
        self.show_focus = tk.BooleanVar(value=True)
        self.timeline_value = tk.IntVar(value=1)
        self.sequence_info = tk.StringVar(value="未加载序列")
        self.frame_info = tk.StringVar(value="未加载图片")
        self.max_sequences = tk.StringVar(value="")
        self.status = tk.StringVar(value="请点击“加载/刷新序列”。")
        self._analysis_running = False
        self._step_config_key = None
        self._step_source_key = None
        self._step_sequence_dirs: list[Path] = []
        self._step_next_index = 0
        self._current_sequence_index = 0
        self._step_summary_rows: list[dict[str, str]] = []
        self._step_frame_rows: list[dict[str, str]] = []
        self._analyzed_sequences: set[str] = set()
        self._current_frame_paths: list[Path] = []
        self._current_frame_index = 0
        self._current_preview_image = None
        self._current_preview_meta: dict[str, Any] = {}
        self._photo_image = None
        self._roi_stat_vars: dict[str, dict[str, Any]] = {}
        self._roi_rect_entry_vars: dict[str, dict[str, Any]] = {}
        self._roi_histogram_labels: dict[str, Any] = {}
        self._roi_histogram_images: dict[str, Any] = {}
        self._settings_window = None
        self._preview_refresh_after_id = None
        self._preview_resize_after_id = None
        self._syncing_roi_rect_inputs = False
        self._last_preview_area_size = (0, 0)
        self._build_ui()
        self._install_setting_traces()
        self.root.after(0, self.load_sequences)

    def _maximize_root(self) -> None:
        self.root.state("zoomed")

    def _build_ui(self) -> None:
        ttk = self.ttk
        self._build_menu()
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main = ttk.Frame(self.root, padding=10)
        main.grid(row=0, column=0, sticky="nsew")
        main.columnconfigure(0, weight=3)
        main.columnconfigure(1, weight=2, minsize=ROI_STATS_PANEL_MIN_WIDTH)
        main.rowconfigure(1, weight=1)

        buttons = ttk.Frame(main)
        buttons.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        self._configure_source_folder_button_style()
        self.tk.Button(
            buttons,
            text="选择文件夹",
            command=self.select_source_folder,
            bg="#22A652",
            fg="white",
            activebackground="#188A43",
            activeforeground="white",
            relief="raised",
            padx=10,
            pady=2,
        ).pack(side="left")
        ttk.Button(buttons, text="加载/刷新序列", command=self.load_sequences).pack(side="left")
        self.analyze_button = ttk.Button(buttons, text="分析当前序列", command=self.run_analysis)
        self.analyze_button.pack(side="left", padx=(8, 0))
        self.next_button = ttk.Button(buttons, text="下一个序列", command=self.next_sequence)
        self.next_button.pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="导出当前序列Excel", command=self.export_current_sequence_excel).pack(side="left", padx=(8, 0))
        self.sequence_selector = tk.StringVar(value="")
        self.sequence_combo = ttk.Combobox(
            buttons,
            textvariable=self.sequence_selector,
            values=[],
            state="readonly",
            width=28,
        )
        self.sequence_combo.pack(side="left", padx=(8, 0))
        self.sequence_combo.bind("<<ComboboxSelected>>", self.on_sequence_selected)
        ttk.Checkbutton(buttons, text="显示ROI1", variable=self.show_roi1, command=self.refresh_preview).pack(side="left", padx=(20, 0))
        ttk.Checkbutton(buttons, text="显示ROI2", variable=self.show_roi2, command=self.refresh_preview).pack(side="left", padx=(8, 0))
        ttk.Checkbutton(buttons, text="显示ROI3", variable=self.show_roi3, command=self.refresh_preview).pack(side="left", padx=(8, 0))
        ttk.Checkbutton(buttons, text="显示ROI4", variable=self.show_roi4, command=self.refresh_preview).pack(side="left", padx=(8, 0))
        ttk.Checkbutton(buttons, text="显示焦点", variable=self.show_focus, command=self.refresh_preview).pack(side="left", padx=(8, 0))
        ttk.Button(buttons, text="保存ROI位置/高亮设置", command=self.save_roi_rect_settings).pack(side="left", padx=(12, 0))
        ttk.Label(buttons, textvariable=self.sequence_info).pack(side="right")
        self.run_button = self.analyze_button

        preview_panel = ttk.Frame(main)
        preview_panel.grid(row=1, column=0, sticky="nsew", pady=(4, 6), padx=(0, 10))
        preview_panel.columnconfigure(0, weight=1)
        preview_panel.rowconfigure(0, weight=1)
        self.image_label = ttk.Label(preview_panel, text="请先加载序列", anchor=PREVIEW_IMAGE_ANCHOR)
        self.image_label.grid(row=0, column=0, sticky="nsew")
        self.image_label.bind("<Configure>", self._on_preview_area_configure)

        self._build_roi_stats_panel(main)

        timeline = ttk.Frame(main)
        timeline.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(4, 2))
        timeline.columnconfigure(1, weight=1)
        ttk.Label(timeline, text="时间轴").grid(row=0, column=0, sticky="w")
        self.timeline = ttk.Scale(timeline, from_=1, to=1, orient="horizontal", command=self.on_timeline_change)
        self.timeline.grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Label(timeline, textvariable=self.frame_info).grid(row=0, column=2, sticky="e")

        ttk.Label(main, textvariable=self.status, wraplength=1200).grid(row=3, column=0, columnspan=2, sticky="ew", pady=(8, 0))

    def _configure_source_folder_button_style(self) -> None:
        style = self.ttk.Style()
        style.configure("Green.TButton", foreground="white", background="#22A652")
        style.map(
            "Green.TButton",
            foreground=[("disabled", "#E6E6E6"), ("active", "white")],
            background=[("disabled", "#8FB99B"), ("active", "#188A43")],
        )

    def select_source_folder(self, ask_directory: Optional[Callable[..., str]] = None) -> None:
        from tkinter import filedialog

        chooser = ask_directory or filedialog.askdirectory
        selected = chooser(title="选择序列根目录", initialdir=self.root_dir.get() or None)
        if not selected:
            return
        self.root_dir.set(selected)
        self.load_sequences()

    def _build_roi_stats_panel(self, parent) -> None:
        ttk = self.ttk
        from tkinter import font as tkfont

        panel = ttk.LabelFrame(parent, text="ROI统计信息", padding=8)
        panel.grid(row=1, column=1, sticky="nsew", pady=(4, 6))
        for column in range(ROI_STATS_CARD_COLUMNS):
            panel.columnconfigure(column, weight=1, uniform="roi_stats")
        stat_font = tkfont.nametofont("TkDefaultFont").copy()
        stat_font.configure(size=ROI_STATS_FONT_SIZE)

        palette = {
            "ROI1": "红色，整帧/背景区域",
            "ROI2": "绿色，焦域ROI",
            "ROI3": "黄色，焦域下ROI",
            "ROI4": "橙色，HEM高亮/底部区域",
        }
        for index, roi_name in enumerate(("ROI1", "ROI2", "ROI3", "ROI4")):
            row, column = roi_stat_card_grid_position(index)
            card = ttk.LabelFrame(panel, text=f"{roi_name}（{palette[roi_name]}）", padding=8)
            card.grid(row=row, column=column, sticky="nsew", padx=(0 if column == 0 else 6, 0), pady=(0, 8))
            card.columnconfigure(1, weight=1)
            card.columnconfigure(3, weight=1)
            card.columnconfigure(5, weight=1)
            card.columnconfigure(7, weight=1)
            panel.rowconfigure(row, weight=1)
            rect_vars = self.roi_rect_vars[roi_name]
            self._roi_rect_entry_vars[roi_name] = rect_vars
            ttk.Label(card, text="形状", font=stat_font).grid(row=0, column=0, sticky="w", padx=(0, 2), pady=(0, 3))
            ttk.Combobox(
                card,
                textvariable=self.roi_shape_vars[roi_name],
                values=tuple(ROI_SHAPE_LABELS.keys()),
                state="readonly",
                width=6,
                font=stat_font,
            ).grid(row=0, column=1, columnspan=3, sticky="ew", padx=(0, 5), pady=(0, 3))
            for col, (label, field_name) in enumerate((("X", "x"), ("Y", "y"), ("宽", "width"), ("高", "height"))):
                ttk.Label(card, text=label, font=stat_font).grid(row=1, column=col * 2, sticky="w", padx=(0, 2), pady=(0, 3))
                ttk.Entry(card, textvariable=rect_vars[field_name], width=6, font=stat_font).grid(
                    row=1,
                    column=col * 2 + 1,
                    sticky="ew",
                    padx=(0, 5),
                    pady=(0, 3),
                )
            histogram_label = ttk.Label(card)
            histogram_label.grid(row=2, column=0, columnspan=8, sticky="ew", pady=(0, 3))
            self._roi_histogram_labels[roi_name] = histogram_label
            values = {field: self.tk.StringVar(value="-") for field, _label in ROI_STAT_DISPLAY_FIELDS}
            self._roi_stat_vars[roi_name] = values
            for field_index, (field, label) in enumerate(ROI_STAT_DISPLAY_FIELDS):
                stat_row, stat_column = roi_stat_value_grid_position(field_index)
                row_index = stat_row + 3
                label_column = stat_column * 2
                value_column = label_column + 1
                ttk.Label(card, text=label, font=stat_font).grid(
                    row=row_index,
                    column=label_column,
                    sticky="w",
                    padx=(0, 2),
                    pady=ROI_STATS_ROW_PADDING,
                )
                ttk.Label(card, textvariable=values[field], font=stat_font).grid(
                    row=row_index,
                    column=value_column,
                    sticky="e",
                    padx=(0, 5),
                    pady=ROI_STATS_ROW_PADDING,
                )

    def _build_menu(self) -> None:
        menu = self.tk.Menu(self.root)
        sequence_menu = self.tk.Menu(menu, tearoff=False)
        sequence_menu.add_command(label="加载/刷新序列", command=self.load_sequences)
        sequence_menu.add_command(label="分析当前序列", command=self.run_analysis)
        sequence_menu.add_command(label="下一个序列", command=self.next_sequence)
        menu.add_cascade(label="序列", menu=sequence_menu)

        settings_menu = self.tk.Menu(menu, tearoff=False)
        settings_menu.add_command(label="参数设置...", command=self.open_settings_dialog)
        settings_menu.add_command(label="加载配置默认值", command=self.load_settings_defaults)
        menu.add_cascade(label="设置", menu=settings_menu)
        self.root.config(menu=menu)

    def open_settings_dialog(self) -> None:
        if self._settings_window is not None and self._settings_window.winfo_exists():
            self._settings_window.lift()
            return

        ttk = self.ttk
        win = self.tk.Toplevel(self.root)
        self._settings_window = win
        win.title("参数设置")
        win.columnconfigure(0, weight=1)
        content = ttk.Frame(win, padding=12)
        content.grid(row=0, column=0, sticky="nsew")
        content.columnconfigure(0, weight=1)

        paths = ttk.LabelFrame(content, text="路径设置", padding=8)
        paths.grid(row=0, column=0, sticky="ew")
        paths.columnconfigure(1, weight=1)
        row = 0
        row = self._path_row(paths, row, "序列根目录", self.root_dir, "dir")
        row = self._path_row(paths, row, "汇总CSV", self.output_csv, "save_csv")
        row = self._path_row(paths, row, "逐帧CSV", self.per_frame_csv, "save_csv")
        row = self._path_row(paths, row, "配置JSON", self.settings_path, "file")
        row = self._path_row(paths, row, "焦点CSV", self.focus_points_csv, "file")
        row = self._path_row(paths, row, "Excel导出目录", self.excel_output_dir, "dir")

        focus = ttk.LabelFrame(content, text="焦点设置", padding=8)
        focus.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        for col in range(4):
            focus.columnconfigure(col, weight=1)
        ttk.Label(focus, text="焦点X").grid(row=0, column=0, sticky="w", pady=4)
        ttk.Entry(focus, textvariable=self.focus_x, width=12).grid(row=0, column=1, sticky="ew", padx=(8, 16), pady=4)
        ttk.Label(focus, text="焦点Y").grid(row=0, column=2, sticky="w", pady=4)
        ttk.Entry(focus, textvariable=self.focus_y, width=12).grid(row=0, column=3, sticky="ew", padx=(8, 0), pady=4)
        self._entry_row(focus, 1, "超声深度mm", self.provider_depth_mm, "焦点y偏移>0时必填")
        self._entry_row(focus, 2, "焦点y偏移mm", self.focus_y_offset_mm, "ROI2锚点按当前算法偏移")

        roi = ttk.LabelFrame(content, text="ROI2扩展参数", padding=8)
        roi.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        for col in range(8):
            roi.columnconfigure(col, weight=1)
        for col, (label, var) in enumerate(
            [
                ("左", self.roi2_left),
                ("右", self.roi2_right),
                ("上", self.roi2_top),
                ("下", self.roi2_bottom),
            ]
        ):
            ttk.Label(roi, text=label).grid(row=0, column=col * 2, sticky="w", padx=(0, 4), pady=4)
            ttk.Entry(roi, textvariable=var, width=8).grid(row=0, column=col * 2 + 1, sticky="ew", padx=(0, 8), pady=4)
        ttk.Label(roi, text="绿框由焦点/偏移锚点和四向扩展共同决定，修改后会自动刷新当前预览。").grid(
            row=1, column=0, columnspan=8, sticky="w", pady=(4, 0)
        )

        roi3 = ttk.LabelFrame(content, text="ROI3扩展参数（黄色，焦域下方区域）", padding=8)
        roi3.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        for col in range(8):
            roi3.columnconfigure(col, weight=1)
        for col, (label, var) in enumerate(
            [
                ("左", self.roi3_left),
                ("右", self.roi3_right),
                ("上", self.roi3_top),
                ("下", self.roi3_bottom),
            ]
        ):
            ttk.Label(roi3, text=label).grid(row=0, column=col * 2, sticky="w", padx=(0, 4), pady=4)
            ttk.Entry(roi3, textvariable=var, width=8).grid(row=0, column=col * 2 + 1, sticky="ew", padx=(0, 8), pady=4)
        ttk.Label(roi3, text="黄框与当前算法 peak_detect.roi3_extension_params 一致，随焦点/偏移锚点联动。").grid(
            row=1, column=0, columnspan=8, sticky="w", pady=(4, 0)
        )

        roi4 = ttk.LabelFrame(content, text="ROI4区域（橙色，高亮候选/底部区域）", padding=8)
        roi4.grid(row=4, column=0, sticky="ew", pady=(8, 0))
        for col in range(8):
            roi4.columnconfigure(col, weight=1)
        for col, (label, var) in enumerate(
            [
                ("X", self.roi4_x),
                ("Y", self.roi4_y),
                ("宽", self.roi4_width),
                ("高", self.roi4_height),
            ]
        ):
            ttk.Label(roi4, text=label).grid(row=0, column=col * 2, sticky="w", padx=(0, 4), pady=4)
            ttk.Entry(roi4, textvariable=var, width=8).grid(row=0, column=col * 2 + 1, sticky="ew", padx=(0, 8), pady=4)
        self._entry_row(roi4, 1, "底部高度比例", self.roi4_bottom_region_ratio, "固定区域留空时生效，当前默认0.3")
        ttk.Label(roi4, text="固定区域优先；X/Y/宽/高全部留空时，按底部高度比例生成ROI4。").grid(
            row=2, column=0, columnspan=8, sticky="w", pady=(4, 0)
        )

        analysis = ttk.LabelFrame(content, text="分析参数", padding=8)
        analysis.grid(row=5, column=0, sticky="ew", pady=(8, 0))
        analysis.columnconfigure(1, weight=1)
        row = self._entry_row(analysis, 0, "差值阈值", self.difference_threshold, "")
        row = self._entry_row(analysis, row, "基准帧序号", self.before_frame_index, "从1开始")
        ttk.Label(analysis, text="治疗后帧选择").grid(row=row, column=0, sticky="w", pady=4)
        strategy = ttk.Combobox(
            analysis,
            textvariable=self.after_strategy,
            values=tuple(AFTER_STRATEGY_LABELS.keys()),
            state="readonly",
            width=16,
        )
        strategy.grid(row=row, column=1, sticky="w", pady=4)
        row += 1
        ttk.Checkbutton(analysis, text="包含 selected_*.png 调试图片", variable=self.include_selected_debug).grid(
            row=row, column=1, sticky="w", pady=4
        )
        row += 1
        row = self._entry_row(analysis, row, "最大序列数", self.max_sequences, "可选，用于小样本试跑")
        ttk.Label(analysis, text="高亮模式").grid(row=row, column=0, sticky="w", pady=4)
        ttk.Combobox(
            analysis,
            textvariable=self.highlight_mode,
            values=tuple(HIGHLIGHT_MODE_LABELS.keys()),
            state="readonly",
            width=16,
        ).grid(row=row, column=1, sticky="w", pady=4)
        row += 1
        row = self._entry_row(analysis, row, "固定灰度阈值", self.highlight_fixed_gray, "0~255，固定灰度值模式生效")
        self._entry_row(analysis, row, "基线倍数", self.highlight_baseline_multiplier, "基线倍数模式生效")

        actions = ttk.Frame(content)
        actions.grid(row=6, column=0, sticky="ew", pady=(10, 0))
        ttk.Button(actions, text="应用并刷新", command=self.apply_settings_changes).pack(side="right")
        ttk.Button(actions, text="关闭", command=win.destroy).pack(side="right", padx=(0, 8))
        win.protocol("WM_DELETE_WINDOW", win.destroy)

    def _entry_row(self, parent, row: int, label: str, var, hint: str) -> int:
        ttk = self.ttk
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Label(parent, text=hint).grid(row=row, column=2, sticky="w", padx=(8, 0), pady=4)
        return row + 1

    def _path_row(self, parent, row: int, label: str, var, mode: str) -> int:
        ttk = self.ttk
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=4)
        ttk.Entry(parent, textvariable=var).grid(row=row, column=1, sticky="ew", pady=4)
        ttk.Button(parent, text="浏览", command=lambda: self.browse_path(var, mode)).grid(row=row, column=2, sticky="ew", padx=(8, 0), pady=4)
        return row + 1

    def browse_path(self, var, mode: str) -> None:
        from tkinter import filedialog

        if mode == "dir":
            selected = filedialog.askdirectory(title="选择序列根目录")
        elif mode == "save_csv":
            selected = filedialog.asksaveasfilename(
                title="选择CSV输出文件",
                defaultextension=".csv",
                filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")],
            )
        else:
            selected = filedialog.askopenfilename(title="选择文件")
        if selected:
            var.set(selected)

    def _focus_point_from_xy(self) -> str:
        x_text = self.focus_x.get().strip()
        y_text = self.focus_y.get().strip()
        if not x_text and not y_text:
            return ""
        try:
            x = float(x_text)
            y = float(y_text)
        except ValueError as exc:
            raise ValueError("焦点X/Y必须是数字") from exc
        return f"PointF({x}, {y})"

    def _sync_focus_point_from_xy(self, strict: bool = False) -> None:
        try:
            self.focus_point.set(self._focus_point_from_xy())
        except Exception:
            if strict:
                raise
            return

    def _source_key_from_state(self, state: GuiState) -> tuple:
        return (
            state.root_dir,
            state.include_selected_debug,
            state.max_sequences,
        )

    def _source_key_from_config(self, config: AnalyzerConfig) -> tuple:
        return (
            str(config.root_dir),
            config.include_selected_debug,
            "" if config.max_sequences is None else str(config.max_sequences),
        )

    def _install_setting_traces(self) -> None:
        live_preview_vars = [
            self.focus_x,
            self.focus_y,
            self.focus_points_csv,
            self.provider_depth_mm,
            self.focus_y_offset_mm,
            self.roi2_left,
            self.roi2_right,
            self.roi2_top,
            self.roi2_bottom,
            self.roi3_left,
            self.roi3_right,
            self.roi3_top,
            self.roi3_bottom,
            self.roi4_x,
            self.roi4_y,
            self.roi4_width,
            self.roi4_height,
            self.roi4_bottom_region_ratio,
        ]
        for roi_name in ROI_NAMES:
            for field_name in ROI_RECT_FIELDS:
                live_preview_vars.append(self.roi_rect_vars[roi_name][field_name])
        live_preview_vars.extend(
            [
                self.highlight_mode,
                self.highlight_fixed_gray,
                self.highlight_baseline_multiplier,
            ]
        )
        source_vars = [
            self.root_dir,
            self.include_selected_debug,
            self.max_sequences,
        ]
        analysis_vars = [
            self.output_csv,
            self.per_frame_csv,
            self.settings_path,
            self.excel_output_dir,
            self.difference_threshold,
            self.before_frame_index,
            self.after_strategy,
        ]

        for var in live_preview_vars:
            var.trace_add("write", self._schedule_preview_refresh)
        for roi_name in ROI_NAMES:
            self.roi_shape_vars[roi_name].trace_add(
                "write",
                lambda *_args, roi_name=roi_name: self._on_roi_shape_changed(roi_name),
            )
        for var in source_vars:
            var.trace_add("write", self._mark_sequences_need_reload)
        for var in analysis_vars:
            var.trace_add("write", self._mark_analysis_settings_changed)

    def _schedule_preview_refresh(self, *_args) -> None:
        if getattr(self, "_syncing_roi_rect_inputs", False):
            return
        self._sync_focus_point_from_xy()
        if not self._current_frame_paths:
            return
        if self._preview_refresh_after_id is not None:
            try:
                self.root.after_cancel(self._preview_refresh_after_id)
            except Exception:
                pass
        self._preview_refresh_after_id = self.root.after(250, self._refresh_preview_after_settings_change)

    def _refresh_preview_after_settings_change(self) -> None:
        self._preview_refresh_after_id = None
        try:
            state = self.current_state()
            config = config_from_gui_state(state)
        except Exception as exc:
            self.status.set(f"参数未生效：{exc}")
            return
        if self._step_source_key != self._source_key_from_config(config):
            self.status.set("序列路径或帧过滤参数已改变，请点击“加载/刷新序列”。")
            return
        self._step_config_key = self._config_key(config)
        self.refresh_preview()

    def _mark_sequences_need_reload(self, *_args) -> None:
        if self._step_sequence_dirs:
            self.status.set("序列路径或帧过滤参数已改变，请点击“加载/刷新序列”。")

    def _mark_analysis_settings_changed(self, *_args) -> None:
        self._sync_focus_point_from_xy()
        if self._step_sequence_dirs:
            self.status.set("分析参数已改变，当前预览可继续查看；分析时会使用新参数。")

    def apply_settings_changes(self) -> None:
        try:
            state = self.current_state()
            config = config_from_gui_state(state)
        except Exception as exc:
            self.status.set(f"参数错误：{exc}")
            return
        if self._step_source_key != self._source_key_from_config(config):
            self.status.set("参数已保存。序列路径或帧过滤参数改变，请点击“加载/刷新序列”。")
            return
        self._step_config_key = self._config_key(config)
        self.refresh_preview()

    def _roi_rect_inputs_from_vars(self) -> dict[str, dict[str, str]]:
        return {
            roi_name: {
                field_name: self.roi_rect_vars[roi_name][field_name].get()
                for field_name in ROI_RECT_FIELDS
            }
            for roi_name in ROI_NAMES
        }

    def _roi_shape_inputs_from_vars(self) -> dict[str, str]:
        return {
            roi_name: ROI_SHAPE_LABELS.get(self.roi_shape_vars[roi_name].get(), self.roi_shape_vars[roi_name].get())
            for roi_name in ROI_NAMES
        }

    def _sync_current_preview_roi_rect_inputs(self, meta: dict[str, Any]) -> None:
        self._syncing_roi_rect_inputs = True
        try:
            for roi_name in ROI_NAMES:
                rect = meta.get(f"{roi_name.lower()}_rect")
                if rect is None:
                    continue
                values = _rect_to_input_values(tuple(int(v) for v in rect))
                for field_name in ROI_RECT_FIELDS:
                    self.roi_rect_vars[roi_name][field_name].set(values[field_name])
        finally:
            self._syncing_roi_rect_inputs = False

    def _fill_roi_rect_input_from_preview_if_blank(self, roi_name: str) -> None:
        values = self.roi_rect_vars[roi_name]
        fields = [str(values[field_name].get()).strip() for field_name in ROI_RECT_FIELDS]
        if not all(not field for field in fields):
            return
        rect = getattr(self, "_current_preview_meta", {}).get(f"{roi_name.lower()}_rect")
        if rect is None:
            return
        input_values = _rect_to_input_values(tuple(int(v) for v in rect))
        for field_name in ROI_RECT_FIELDS:
            values[field_name].set(input_values[field_name])

    def _on_roi_shape_changed(self, roi_name: str) -> None:
        if roi_name not in ROI_NAMES:
            raise ValueError(f"unknown ROI name: {roi_name}")
        self._fill_roi_rect_input_from_preview_if_blank(roi_name)
        self._schedule_preview_refresh()

    def _apply_roi_definitions_to_vars(self, definitions: dict[str, dict[str, Any]]) -> None:
        for roi_name in ROI_NAMES:
            definition = definitions.get(roi_name)
            if not definition:
                continue
            values = _rect_to_input_values(definition.get("rect"))
            for field_name in ROI_RECT_FIELDS:
                self.roi_rect_vars[roi_name][field_name].set(values[field_name])
            shape = _normalize_roi_shape(definition.get("shape", "rectangle"), roi_name)
            self.roi_shape_vars[roi_name].set(ROI_SHAPE_DISPLAY[shape])

    def save_roi_rect_settings(self) -> None:
        from tkinter import messagebox

        try:
            state = self.current_state()
            settings_path = _optional_path(state.settings_path)
            if settings_path is None:
                raise ValueError("请先设置配置JSON路径")
            definitions = _roi_definitions_from_inputs_or_preview(
                state.roi_rect_inputs,
                state.roi_shape_inputs,
                getattr(self, "_current_preview_meta", {}),
            )
            if not definitions:
                raise ValueError("请先加载序列并显示ROI后再保存")
            self._apply_roi_definitions_to_vars(definitions)
            state = self.current_state()
            definitions = _roi_definitions_from_inputs(state.roi_rect_inputs, state.roi_shape_inputs)
            highlight_rule = _normalize_highlight_rule(
                {
                    "mode": state.highlight_mode,
                    "fixed_gray": state.highlight_fixed_gray,
                    "baseline_multiplier": state.highlight_baseline_multiplier,
                }
            )
            save_visual_analysis_settings(settings_path, definitions, highlight_rule, Path(state.excel_output_dir))
            config = config_from_gui_state(state)
        except Exception as exc:
            self.status.set(f"保存ROI/高亮设置失败：{exc}")
            messagebox.showerror("保存ROI/高亮设置失败", str(exc))
            return
        self._step_config_key = self._config_key(config)
        self.refresh_preview()
        self.status.set(f"ROI/高亮设置已保存到 {settings_path}，当前预览已刷新。")

    def current_state(self) -> GuiState:
        self._sync_focus_point_from_xy(strict=True)
        return GuiState(
            root_dir=self.root_dir.get(),
            output_csv=self.output_csv.get(),
            per_frame_csv=self.per_frame_csv.get(),
            settings_path=self.settings_path.get(),
            focus_point=self.focus_point.get(),
            focus_points_csv=self.focus_points_csv.get(),
            provider_depth_mm=self.provider_depth_mm.get(),
            focus_y_offset_mm=self.focus_y_offset_mm.get(),
            roi2_left=self.roi2_left.get(),
            roi2_right=self.roi2_right.get(),
            roi2_top=self.roi2_top.get(),
            roi2_bottom=self.roi2_bottom.get(),
            difference_threshold=self.difference_threshold.get(),
            before_frame_index=self.before_frame_index.get(),
            after_strategy=AFTER_STRATEGY_LABELS.get(self.after_strategy.get(), self.after_strategy.get()),
            include_selected_debug=self.include_selected_debug.get(),
            max_sequences=self.max_sequences.get(),
            roi3_left=self.roi3_left.get(),
            roi3_right=self.roi3_right.get(),
            roi3_top=self.roi3_top.get(),
            roi3_bottom=self.roi3_bottom.get(),
            roi4_x=self.roi4_x.get(),
            roi4_y=self.roi4_y.get(),
            roi4_width=self.roi4_width.get(),
            roi4_height=self.roi4_height.get(),
            roi4_bottom_region_ratio=self.roi4_bottom_region_ratio.get(),
            roi_rect_inputs=self._roi_rect_inputs_from_vars(),
            roi_shape_inputs=self._roi_shape_inputs_from_vars(),
            highlight_mode=HIGHLIGHT_MODE_LABELS.get(self.highlight_mode.get(), self.highlight_mode.get()),
            highlight_fixed_gray=self.highlight_fixed_gray.get(),
            highlight_baseline_multiplier=self.highlight_baseline_multiplier.get(),
            excel_output_dir=self.excel_output_dir.get(),
        )

    def load_settings_defaults(self) -> None:
        from tkinter import messagebox

        try:
            settings = _load_settings(Path(self.settings_path.get().strip()))
            roi2 = _settings_roi2_params(settings)
            self.roi2_left.set(str(roi2["left"]))
            self.roi2_right.set(str(roi2["right"]))
            self.roi2_top.set(str(roi2["top"]))
            self.roi2_bottom.set(str(roi2["bottom"]))
            roi3 = _settings_roi3_params(settings)
            self.roi3_left.set(str(roi3["left"]))
            self.roi3_right.set(str(roi3["right"]))
            self.roi3_top.set(str(roi3["top"]))
            self.roi3_bottom.set(str(roi3["bottom"]))
            roi4_rect = _settings_roi4_rect(settings)
            if roi4_rect is None:
                self.roi4_x.set("")
                self.roi4_y.set("")
                self.roi4_width.set("")
                self.roi4_height.set("")
                roi4_ratio = _settings_roi4_bottom_region_ratio(settings)
                self.roi4_bottom_region_ratio.set("" if roi4_ratio is None else str(roi4_ratio))
            else:
                self.roi4_x.set(str(roi4_rect[0]))
                self.roi4_y.set(str(roi4_rect[1]))
                self.roi4_width.set(str(roi4_rect[2] - roi4_rect[0]))
                self.roi4_height.set(str(roi4_rect[3] - roi4_rect[1]))
                self.roi4_bottom_region_ratio.set("")
            self.difference_threshold.set(str(_settings_difference_threshold(settings)))
            self.focus_y_offset_mm.set(str(_settings_focus_y_offset(settings)))
            roi_definitions = _settings_roi_definitions(settings)
            for roi_name in ROI_NAMES:
                definition = roi_definitions.get(roi_name, {})
                values = _rect_to_input_values(definition.get("rect"))
                for field_name in ROI_RECT_FIELDS:
                    self.roi_rect_vars[roi_name][field_name].set(values[field_name])
                self.roi_shape_vars[roi_name].set(ROI_SHAPE_DISPLAY[definition.get("shape", "rectangle")])
            highlight_rule = _settings_highlight_rule(settings)
            self.highlight_mode.set(HIGHLIGHT_MODE_DISPLAY[highlight_rule["mode"]])
            self.highlight_fixed_gray.set(str(highlight_rule["fixed_gray"]))
            self.highlight_baseline_multiplier.set(str(highlight_rule["baseline_multiplier"]))
            self.excel_output_dir.set(str(_settings_excel_output_dir(settings)))
            self.status.set("配置默认值已加载。")
        except Exception as exc:
            self.status.set(f"加载配置失败：{exc}")
            messagebox.showerror("加载配置失败", str(exc))

    def _config_key(self, config: AnalyzerConfig) -> tuple:
        return (
            str(config.root_dir),
            str(config.output_csv),
            str(config.per_frame_csv) if config.per_frame_csv is not None else "",
            str(config.settings_path) if config.settings_path is not None else "",
            str(config.focus_point),
            str(config.focus_points_csv) if config.focus_points_csv is not None else "",
            config.provider_depth_mm,
            config.focus_y_offset_mm,
            tuple(sorted((str(k), int(v)) for k, v in config.roi2_extension_params.items())),
            tuple(sorted((str(k), int(v)) for k, v in config.roi3_extension_params.items())),
            config.roi4_rect,
            config.roi4_bottom_region_ratio,
            tuple((roi_name, tuple(rect)) for roi_name, rect in sorted(config.roi_rect_overrides.items())),
            tuple(
                (roi_name, config.roi_definitions.get(roi_name, {}).get("shape", "rectangle"))
                for roi_name in ROI_NAMES
            ),
            tuple(sorted((str(k), str(v)) for k, v in _normalize_highlight_rule(config.highlight_rule).items())),
            str(config.excel_output_dir),
            config.difference_threshold,
            config.before_frame_index,
            config.after_strategy,
            config.include_selected_debug,
            config.max_sequences,
        )

    def _reset_step_state(self, config: AnalyzerConfig) -> None:
        self._step_config_key = self._config_key(config)
        self._step_source_key = self._source_key_from_config(config)
        self._step_sequence_dirs = list_sequences(config.root_dir, config.max_sequences)
        self._step_next_index = 0
        self._current_sequence_index = 0
        self._step_summary_rows = []
        self._step_frame_rows = []
        self._analyzed_sequences = set()
        self._current_frame_paths = []
        self._current_frame_index = 0
        self._sync_sequence_selector_values()

    def _current_sequence_dir(self) -> Optional[Path]:
        if not self._step_sequence_dirs:
            return None
        if self._current_sequence_index < 0 or self._current_sequence_index >= len(self._step_sequence_dirs):
            return None
        return self._step_sequence_dirs[self._current_sequence_index]

    def _sync_sequence_selector_values(self) -> None:
        if not hasattr(self, "sequence_selector"):
            return
        values = [sequence_dir.name for sequence_dir in self._step_sequence_dirs]
        if hasattr(self, "sequence_combo"):
            self.sequence_combo.configure(values=values)
        sequence_dir = self._current_sequence_dir()
        self.sequence_selector.set(sequence_dir.name if sequence_dir is not None else "")

    def _set_sequence_status(self) -> None:
        sequence_dir = self._current_sequence_dir()
        total = len(self._step_sequence_dirs)
        if sequence_dir is None:
            self.sequence_info.set("未加载序列")
            self._sync_sequence_selector_values()
            return
        self.sequence_info.set(f"序列 {self._current_sequence_index + 1}/{total}: {sequence_dir.name}")
        self._sync_sequence_selector_values()

    def _load_current_frame_paths(self, config: AnalyzerConfig) -> None:
        sequence_dir = self._current_sequence_dir()
        if sequence_dir is None:
            self._current_frame_paths = []
            self._current_frame_index = 0
            return
        frame_paths = list_frame_paths(sequence_dir, config.include_selected_debug)
        if not frame_paths:
            raise ValueError(f"当前序列没有可显示的PNG帧：{sequence_dir}")
        self._current_frame_paths = frame_paths
        self._current_frame_index = 0
        self.timeline.configure(from_=1, to=len(frame_paths))
        self.timeline.set(1)

    def select_sequence_by_name(self, sequence_name: str, config: Optional[AnalyzerConfig] = None) -> None:
        target = sequence_name.strip()
        if not target:
            return
        for index, sequence_dir in enumerate(self._step_sequence_dirs):
            if sequence_dir.name == target:
                if config is None:
                    config = config_from_gui_state(self.current_state())
                self._current_sequence_index = index
                self._step_next_index = index
                self._load_current_frame_paths(config)
                self._set_sequence_status()
                self.refresh_preview()
                return
        raise ValueError(f"未找到序列：{target}")

    def on_sequence_selected(self, _event=None) -> None:
        from tkinter import messagebox

        try:
            config = config_from_gui_state(self.current_state())
            if self._step_source_key != self._source_key_from_config(config):
                self._reset_step_state(config)
            elif self._config_key(config) != self._step_config_key:
                self._step_config_key = self._config_key(config)
            self.select_sequence_by_name(self.sequence_selector.get(), config)
        except Exception as exc:
            self.status.set(f"切换序列失败：{exc}")
            messagebox.showerror("切换序列失败", str(exc))

    def _display_preview_image(self, image) -> None:
        from PIL import ImageTk

        self._current_preview_image = image
        self._photo_image = ImageTk.PhotoImage(image)
        self.image_label.configure(image=self._photo_image, text="")

    def _update_roi_stats_panel(self, meta: dict[str, Any]) -> None:
        from PIL import ImageTk

        stat_vars = getattr(self, "_roi_stat_vars", {})
        stats_by_roi = meta.get("roi_stats", {})
        for roi_name, values in stat_vars.items():
            stats = stats_by_roi.get(roi_name) or {}
            width = stats.get("width") or ""
            height = stats.get("height") or ""
            size = f"{width} × {height}" if width and height else "-"
            histogram = stats.get("histogram", [0] * 256)
            if roi_name in getattr(self, "_roi_histogram_labels", {}):
                image = render_roi_histogram_image(histogram)
                photo = ImageTk.PhotoImage(image)
                self._roi_histogram_images[roi_name] = photo
                self._roi_histogram_labels[roi_name].configure(image=photo)
            for field, var in values.items():
                if field == "size":
                    var.set(_format_roi_stat_display_value(size))
                else:
                    var.set(_format_roi_stat_display_value(stats.get(field)))

    def _preview_max_size(self) -> tuple[int, int]:
        try:
            width = int(self.image_label.winfo_width())
            height = int(self.image_label.winfo_height())
        except Exception:
            width, height = PREVIEW_MAX_SIZE
        if width < 100 or height < 100:
            try:
                width = max(PREVIEW_MAX_SIZE[0], int(self.root.winfo_width()) - 24)
                height = max(PREVIEW_MAX_SIZE[1], int(self.root.winfo_height()) - 130)
            except Exception:
                width, height = PREVIEW_MAX_SIZE
        return max(100, width - 8), max(100, height - 8)

    def _on_preview_area_configure(self, event) -> None:
        if not self._current_frame_paths:
            return
        new_size = (int(event.width), int(event.height))
        old_width, old_height = self._last_preview_area_size
        if abs(new_size[0] - old_width) < 4 and abs(new_size[1] - old_height) < 4:
            return
        self._last_preview_area_size = new_size
        if self._preview_resize_after_id is not None:
            try:
                self.root.after_cancel(self._preview_resize_after_id)
            except Exception:
                pass
        self._preview_resize_after_id = self.root.after(120, self._refresh_preview_after_resize)

    def _refresh_preview_after_resize(self) -> None:
        self._preview_resize_after_id = None
        self.refresh_preview()

    def refresh_preview(self) -> None:
        try:
            config = config_from_gui_state(self.current_state())
            sequence_dir = self._current_sequence_dir()
            if sequence_dir is None:
                return
            if not self._current_frame_paths:
                self._load_current_frame_paths(config)
            frame_path = self._current_frame_paths[self._current_frame_index]
            baseline_frame_path = None
            baseline_index = int(config.before_frame_index) - 1
            if 0 <= baseline_index < len(self._current_frame_paths):
                baseline_frame_path = self._current_frame_paths[baseline_index]
            focus_points = load_focus_points_csv(config.focus_points_csv)
            image, meta = render_sequence_preview_image(
                frame_path,
                sequence_dir.name,
                config,
                focus_points,
                bool(self.show_roi2.get()),
                bool(self.show_focus.get()),
                max_size=self._preview_max_size(),
                show_roi1=bool(self.show_roi1.get()),
                show_roi3=bool(self.show_roi3.get()),
                show_roi4=bool(self.show_roi4.get()),
                baseline_frame_path=baseline_frame_path,
            )
            self._current_preview_meta = meta
            self._sync_current_preview_roi_rect_inputs(meta)
            self._display_preview_image(image)
            self._update_roi_stats_panel(meta)
            self._set_sequence_status()
            self.frame_info.set(f"帧 {self._current_frame_index + 1}/{len(self._current_frame_paths)}: {frame_path.name}")
            self.status.set(
                f"已显示 {sequence_dir.name} 的第 {self._current_frame_index + 1} 帧。"
                f" ROI1={_fmt_rect(meta['roi1_rect'])}"
                f" ROI2={_fmt_rect(meta['roi2_rect'])}"
                f" ROI3={_fmt_rect(meta['roi3_rect'])}"
                f" ROI4={_fmt_rect(meta['roi4_rect'])}"
                f" 焦点={_fmt_point(meta['focus_anchor'])}"
            )
        except Exception as exc:
            self.status.set(f"刷新预览失败：{exc}")

    def load_sequences(self) -> None:
        from tkinter import messagebox

        try:
            config = config_from_gui_state(self.current_state())
            self._reset_step_state(config)
            if not self._step_sequence_dirs:
                self.status.set(f"未找到子文件夹：{config.root_dir}")
                self.image_label.configure(image="", text="未找到序列")
                return
            self._load_current_frame_paths(config)
            self._set_sequence_status()
            self.refresh_preview()
        except Exception as exc:
            self.status.set(f"加载序列失败：{exc}")
            messagebox.showerror("加载序列失败", str(exc))

    def on_timeline_change(self, value) -> None:
        if not self._current_frame_paths:
            return
        try:
            frame_index = int(round(float(value))) - 1
        except Exception:
            return
        frame_index = max(0, min(frame_index, len(self._current_frame_paths) - 1))
        if frame_index == self._current_frame_index:
            return
        self._current_frame_index = frame_index
        self.refresh_preview()

    def next_sequence(self) -> None:
        from tkinter import messagebox

        try:
            config = config_from_gui_state(self.current_state())
            if self._step_source_key != self._source_key_from_config(config):
                self._reset_step_state(config)
            elif self._config_key(config) != self._step_config_key:
                self._step_config_key = self._config_key(config)
            if not self._step_sequence_dirs:
                self.load_sequences()
                return
            if self._current_sequence_index + 1 >= len(self._step_sequence_dirs):
                self.status.set("已经是最后一个序列。")
                messagebox.showinfo("序列已结束", "已经是最后一个序列。")
                return
            self._current_sequence_index += 1
            self._step_next_index = self._current_sequence_index
            self._load_current_frame_paths(config)
            self._set_sequence_status()
            self.refresh_preview()
        except Exception as exc:
            self.status.set(f"切换序列失败：{exc}")
            messagebox.showerror("切换序列失败", str(exc))

    def export_current_sequence_excel(
        self,
        ask_directory: Optional[Callable[..., str]] = None,
        show_info: Optional[Callable[..., None]] = None,
        show_error: Optional[Callable[..., None]] = None,
    ) -> None:
        from tkinter import filedialog, messagebox

        try:
            config = config_from_gui_state(self.current_state())
            sequence_dir = self._current_sequence_dir()
            if sequence_dir is None:
                raise ValueError("未选择当前序列")
            if not self._current_frame_paths:
                self._load_current_frame_paths(config)
            chooser = ask_directory or filedialog.askdirectory
            selected_dir = chooser(title="选择Excel导出目录", initialdir=str(config.excel_output_dir))
            if not selected_dir:
                return
            output_dir = Path(selected_dir)
            self.excel_output_dir.set(str(output_dir))
            config = config_from_gui_state(self.current_state())
            focus_points = load_focus_points_csv(config.focus_points_csv)
            output_path = output_dir / f"{sequence_dir.name}.xlsx"
            saved_path = export_sequence_excel(sequence_dir, config, focus_points, output_path)
        except Exception as exc:
            self.status.set(f"导出Excel失败：{exc}")
            (show_error or messagebox.showerror)("导出Excel失败", str(exc))
            return
        self.status.set(f"当前序列Excel已导出：{saved_path}")
        (show_info or messagebox.showinfo)("导出Excel完成", f"已导出：{saved_path}")

    def run_analysis(self) -> None:
        from tkinter import messagebox

        try:
            config = config_from_gui_state(self.current_state())
        except Exception as exc:
            self.status.set(f"分析失败：{exc}")
            messagebox.showerror("分析失败", str(exc))
            return

        if self._analysis_running:
            self.status.set("分析正在运行，请稍候。")
            return

        config_key = self._config_key(config)
        if config_key != self._step_config_key:
            try:
                if self._step_source_key != self._source_key_from_config(config):
                    self._reset_step_state(config)
                    self._load_current_frame_paths(config)
                else:
                    self._step_config_key = config_key
                self.refresh_preview()
            except Exception as exc:
                self.status.set(f"分析失败：{exc}")
                messagebox.showerror("分析失败", str(exc))
                return

        total = len(self._step_sequence_dirs)
        if total == 0:
            self.status.set(f"未找到子文件夹：{config.root_dir}")
            return
        sequence_dir = self._current_sequence_dir()
        if sequence_dir is None:
            self.status.set("未选择当前序列。")
            return

        self._analysis_running = True
        self.run_button.state(["disabled"])
        self.status.set(f"正在分析当前序列 {self._current_sequence_index + 1}/{total}: {sequence_dir.name}")

        def finish_success(row: dict[str, str], frame_rows: list[dict[str, str]]) -> None:
            sequence_name = row["sequence"]
            if sequence_name in self._analyzed_sequences:
                self._step_summary_rows = [saved for saved in self._step_summary_rows if saved["sequence"] != sequence_name]
                self._step_frame_rows = [saved for saved in self._step_frame_rows if saved["sequence"] != sequence_name]
            self._step_summary_rows.append(row)
            self._step_frame_rows.extend(frame_rows)
            self._analyzed_sequences.add(sequence_name)
            self._step_next_index = self._current_sequence_index + 1
            write_csv(config.output_csv, SUMMARY_FIELDS, self._step_summary_rows)
            if config.per_frame_csv is not None:
                write_csv(config.per_frame_csv, FRAME_FIELDS, self._step_frame_rows)
            self._analysis_running = False
            self.run_button.state(["!disabled"])
            color_text = "绿" if row["roi2_color"] == "green" else "红"
            self.status.set(
                f"已分析当前序列 {self._current_sequence_index + 1}/{total}: {sequence_dir.name}。"
                f" 结果={color_text} ROI2差值={row['roi2_diff']} 汇总={config.output_csv}"
            )
            messagebox.showinfo(
                "当前序列分析完成",
                f"序列：{sequence_dir.name}\n"
                f"结果：{color_text}\n"
                f"ROI2差值：{row['roi2_diff']}\n"
                f"汇总CSV：{config.output_csv}",
            )

        def finish_error(exc: Exception) -> None:
            self._analysis_running = False
            self.run_button.state(["!disabled"])
            self.status.set(f"分析失败：{exc}")
            messagebox.showerror("分析失败", str(exc))

        def worker() -> None:
            try:
                focus_points = load_focus_points_csv(config.focus_points_csv)
                row, frame_rows = analyze_sequence(sequence_dir, config, focus_points)
            except Exception as exc:
                self.root.after(0, finish_error, exc)
                return
            self.root.after(0, finish_success, row, frame_rows)

        threading.Thread(target=worker, daemon=True).start()


def launch_gui() -> int:
    import tkinter as tk

    root = tk.Tk()
    HemRoi2BatchAnalyzerGui(root)
    root.mainloop()
    return 0


def main(argv: Optional[list[str]] = None) -> int:
    if argv is None:
        argv = sys.argv[1:]
    if not argv or argv == ["--gui"]:
        return launch_gui()
    config = build_config_from_args(argv)
    if "--gui" in argv:
        return launch_gui()
    rows = analyze_root(config)
    print(f"analyzed_sequences={len(rows)}")
    print(f"summary_csv={config.output_csv}")
    if config.per_frame_csv is not None:
        print(f"per_frame_csv={config.per_frame_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
