import csv
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path

import numpy as np
from PIL import Image

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "resource" / "pywrapper"))
import api_server

from tools import hem_roi2_batch_analyzer as analyzer


def write_frame(path: Path, value: int, size=(20, 20)) -> None:
    arr = np.full((size[1], size[0], 3), value, dtype=np.uint8)
    Image.fromarray(arr).save(path)


class HemRoi2BatchAnalyzerTests(unittest.TestCase):
    def test_default_focus_point_matches_current_algorithm_focus(self):
        self.assertEqual(
            analyzer.DEFAULT_FOCUS_POINT,
            "PointF(299.2863464355469, 285.9410705566406)",
        )
        self.assertEqual(
            analyzer._parse_focus_point_value(analyzer.DEFAULT_FOCUS_POINT),
            (299, 285),
        )

    def test_analyze_sequence_uses_current_roi2_rect_and_returns_green(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "114806_20260614_000000_862"
            seq.mkdir()
            write_frame(seq / "00001_2026-06-14_00-00-00.000_frame.png", 10)
            write_frame(seq / "00002_2026-06-14_00-00-00.070_frame.png", 20)
            write_frame(seq / "00003_2026-06-14_00-00-00.140_frame.png", 12)

            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(10, 10)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 2, "right": 2, "top": 3, "bottom": 3},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="roi2_peak",
                include_selected_debug=False,
                max_sequences=None,
            )

            row, frames = analyzer.analyze_sequence(seq, cfg, {})

            expected_rect = api_server.compute_roi_region(
                (20, 20),
                (10, 10),
                {"left": 2, "right": 2, "top": 3, "bottom": 3},
            )
            self.assertEqual(row["focus_anchor"], "10,10")
            self.assertEqual(row["offset_anchor"], "10,10")
            self.assertEqual(row["roi2_rect"], ",".join(str(v) for v in expected_rect))
            self.assertEqual(row["before_mean"], "10.000000")
            self.assertEqual(row["after_mean"], "20.000000")
            self.assertEqual(row["roi2_diff"], "10.000000")
            self.assertEqual(row["roi2_color"], "green")
            self.assertEqual(row["after_frame_index"], "2")
            self.assertEqual(len(frames), 3)

    def test_write_outputs_creates_summary_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            write_frame(seq / "00001_2026-06-14_00-00-00.000_frame.png", 10)
            write_frame(seq / "00002_2026-06-14_00-00-00.070_frame.png", 18)
            write_frame(seq / "after_roi2.png", 255, size=(4, 4))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=root / "frames.csv",
                settings_path=None,
                focus_point=[10, 10],
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 2, "right": 2, "top": 3, "bottom": 3},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )

            rows = analyzer.analyze_root(cfg)

            self.assertEqual(len(rows), 1)
            with cfg.output_csv.open("r", encoding="utf-8-sig", newline="") as f:
                saved = list(csv.DictReader(f))
            self.assertEqual(saved[0]["sequence"], "seq")
            self.assertEqual(saved[0]["frame_count"], "2")
            self.assertEqual(saved[0]["roi2_color"], "green")
            self.assertTrue(cfg.per_frame_csv.exists())

    def test_missing_focus_point_fails_fast(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            write_frame(seq / "00001_2026-06-14_00-00-00.000_frame.png", 10)
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point=None,
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 2, "right": 2, "top": 3, "bottom": 3},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )

            with self.assertRaisesRegex(ValueError, "focus_point is required"):
                analyzer.analyze_sequence(seq, cfg, {})

    def test_focus_y_offset_requires_depth(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            write_frame(seq / "00001_2026-06-14_00-00-00.000_frame.png", 10, size=(200, 200))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(100, 100)",
                focus_points_csv=None,
                provider_depth_mm=None,
                focus_y_offset_mm=2.5,
                roi2_extension_params={"left": 5, "right": 5, "top": 6, "bottom": 6},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )

            with self.assertRaisesRegex(ValueError, "provider ultrasound depth is required"):
                analyzer.analyze_sequence(seq, cfg, {})

    def test_render_sequence_preview_draws_roi2_and_focus_toggles(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            frame = seq / "00001_2026-06-14_00-00-00.000_frame.png"
            write_frame(frame, 10, size=(200, 200))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(100, 100)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 5, "right": 5, "top": 5, "bottom": 5},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )

            hidden, hidden_meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                False,
                False,
                max_size=(200, 200),
            )
            visible, visible_meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                True,
                True,
                max_size=(200, 200),
            )

            self.assertEqual(hidden_meta["focus_anchor"], (100, 100))
            self.assertEqual(hidden_meta["roi2_rect"], (95, 95, 105, 105))
            self.assertEqual(visible_meta["roi2_rect"], hidden_meta["roi2_rect"])
            self.assertNotEqual(visible.getpixel((95, 95)), hidden.getpixel((95, 95)))
            self.assertNotEqual(visible.getpixel((100, 100)), hidden.getpixel((100, 100)))

    def test_render_sequence_preview_draws_algorithm_roi1_roi3_roi4(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            frame = seq / "00001_2026-06-14_00-00-00.000_frame.png"
            write_frame(frame, 10, size=(200, 200))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(100, 100)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 5, "right": 5, "top": 5, "bottom": 5},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
                roi3_extension_params={"left": 10, "right": 10, "top": 20, "bottom": 30},
                roi4_bottom_region_ratio=0.25,
            )

            hidden, _hidden_meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                False,
                False,
                max_size=(200, 200),
                show_roi1=False,
                show_roi3=False,
                show_roi4=False,
            )
            visible, visible_meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                True,
                False,
                max_size=(200, 200),
                show_roi1=True,
                show_roi3=True,
                show_roi4=True,
            )

            self.assertEqual(visible_meta["roi1_rect"], (0, 0, 200, 200))
            self.assertEqual(visible_meta["roi2_rect"], (95, 95, 105, 105))
            self.assertEqual(visible_meta["roi3_rect"], (90, 80, 110, 130))
            self.assertEqual(visible_meta["roi4_rect"], (0, 150, 200, 200))
            self.assertNotEqual(visible.getpixel((1, 1)), hidden.getpixel((1, 1)))
            self.assertNotEqual(visible.getpixel((5, 150)), hidden.getpixel((5, 150)))
            self.assertNotEqual(visible.getpixel((90, 80)), hidden.getpixel((90, 80)))
            self.assertNotEqual(visible.getpixel((95, 95)), hidden.getpixel((95, 95)))

    def test_render_sequence_preview_returns_roi_stats_for_sidebar(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            frame = seq / "00001_2026-06-14_00-00-00.000_frame.png"
            write_frame(frame, 10, size=(200, 200))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(100, 100)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 5, "right": 5, "top": 5, "bottom": 5},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
                roi3_extension_params={"left": 10, "right": 10, "top": 20, "bottom": 30},
                roi4_bottom_region_ratio=0.25,
            )

            _image, meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                True,
                False,
                max_size=(200, 200),
                show_roi1=True,
                show_roi3=True,
                show_roi4=True,
            )

            stats = meta["roi_stats"]
            self.assertEqual(stats["ROI1"]["rect"], "0,0,200,200")
            self.assertEqual(stats["ROI1"]["area"], "40000")
            self.assertEqual(stats["ROI1"]["mean"], "10.000000")
            self.assertEqual(stats["ROI2"]["width"], "10")
            self.assertEqual(stats["ROI2"]["height"], "10")
            self.assertEqual(stats["ROI3"]["rect"], "90,80,110,130")
            self.assertEqual(stats["ROI4"]["rect"], "0,150,200,200")

    def test_roi_stats_for_frame_reports_ppt_gray_distribution_metrics(self):
        values = np.arange(16, dtype=np.uint8).reshape((4, 4))
        frame = np.repeat(values[:, :, None], 3, axis=2)

        stats = analyzer.roi_stats_for_frame(frame, (0, 0, 4, 4))

        self.assertEqual(stats["rect"], "0,0,4,4")
        self.assertEqual(stats["area"], "16")
        self.assertEqual(stats["mean"], "7.500000")
        self.assertEqual(stats["density"], "0.029412")
        self.assertEqual(stats["std"], "4.609772")
        self.assertEqual(stats["median"], "7.500000")
        self.assertEqual(stats["median_abs_deviation"], "4.000000")
        self.assertEqual(stats["p10"], "1.500000")
        self.assertEqual(stats["p90"], "13.500000")
        self.assertEqual(stats["threshold"], "8.625000")
        self.assertEqual(stats["highlight_count"], "7")
        self.assertEqual(stats["highlight_area"], "7")
        self.assertEqual(stats["highlight_ratio"], "0.437500")
        self.assertEqual(stats["highlight_std"], "2.000000")

    def test_roi_stats_for_frame_reports_baseline_and_hem_area_metrics(self):
        baseline_values = np.arange(16, dtype=np.uint8).reshape((4, 4))
        current_values = baseline_values + 20
        baseline_frame = np.repeat(baseline_values[:, :, None], 3, axis=2)
        current_frame = np.repeat(current_values[:, :, None], 3, axis=2)
        baseline_stats = analyzer.roi_stats_for_frame(baseline_frame, (0, 0, 4, 4))

        stats = analyzer.roi_stats_for_frame(current_frame, (0, 0, 4, 4), baseline_stats)

        self.assertEqual(stats["threshold"], "8.625000")
        self.assertEqual(stats["highlight_count"], "16")
        self.assertEqual(stats["hem_z_count"], "14")
        self.assertEqual(stats["hem_z_area"], "14")
        self.assertEqual(stats["mean_delta"], "20.000000")
        self.assertEqual(stats["mean_delta_pct"], "2.666667")
        self.assertEqual(stats["std_delta"], "0.000000")
        self.assertEqual(stats["median_delta"], "20.000000")
        self.assertEqual(stats["highlight_area_delta"], "9")

    def test_roi_histogram_counts_rectangle_gray_bins(self):
        values = np.asarray([[0, 1], [1, 255]], dtype=np.uint8)
        frame = np.repeat(values[:, :, None], 3, axis=2)

        histogram = analyzer.roi_gray_histogram(frame, (0, 0, 2, 2), shape="rectangle")
        stats = analyzer.roi_stats_for_frame(frame, (0, 0, 2, 2), shape="rectangle")

        self.assertEqual(len(histogram), 256)
        self.assertEqual(histogram[0], 1)
        self.assertEqual(histogram[1], 2)
        self.assertEqual(histogram[255], 1)
        self.assertEqual(sum(histogram), 4)
        self.assertEqual(stats["histogram"][1], 2)

    def test_ellipse_roi_uses_masked_pixels_for_stats_and_histogram(self):
        values = np.arange(25, dtype=np.uint8).reshape((5, 5))
        frame = np.repeat(values[:, :, None], 3, axis=2)

        rectangle = analyzer.roi_stats_for_frame(frame, (0, 0, 5, 5), shape="rectangle")
        ellipse = analyzer.roi_stats_for_frame(frame, (0, 0, 5, 5), shape="ellipse")

        self.assertEqual(rectangle["area"], "25")
        self.assertLess(int(ellipse["area"]), 25)
        self.assertEqual(sum(ellipse["histogram"]), int(ellipse["area"]))

    def test_fixed_gray_highlight_rule_counts_threshold_pixels(self):
        values = np.asarray([[9, 10, 11]], dtype=np.uint8)
        frame = np.repeat(values[:, :, None], 3, axis=2)

        stats = analyzer.roi_stats_for_frame(
            frame,
            (0, 0, 3, 1),
            highlight_rule={"mode": "fixed_gray", "fixed_gray": 10, "baseline_multiplier": 1.15},
        )

        self.assertEqual(stats["threshold"], "10.000000")
        self.assertEqual(stats["highlight_count"], "2")

    def test_baseline_multiplier_highlight_rule_uses_baseline_mean(self):
        baseline = np.full((1, 3), 10, dtype=np.uint8)
        current = np.asarray([[14, 15, 16]], dtype=np.uint8)
        baseline_frame = np.repeat(baseline[:, :, None], 3, axis=2)
        current_frame = np.repeat(current[:, :, None], 3, axis=2)
        rule = {"mode": "baseline_multiplier", "fixed_gray": 93, "baseline_multiplier": 1.5}
        baseline_stats = analyzer.roi_stats_for_frame(baseline_frame, (0, 0, 3, 1), highlight_rule=rule)

        stats = analyzer.roi_stats_for_frame(
            current_frame,
            (0, 0, 3, 1),
            baseline_stats=baseline_stats,
            highlight_rule=rule,
        )

        self.assertEqual(stats["threshold"], "15.000000")
        self.assertEqual(stats["highlight_count"], "2")

    def test_render_sequence_preview_scales_up_to_available_area(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq"
            seq.mkdir()
            frame = seq / "00001_2026-06-14_00-00-00.000_frame.png"
            write_frame(frame, 10, size=(200, 100))
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(100, 50)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 5, "right": 5, "top": 5, "bottom": 5},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )

            image, meta = analyzer.render_sequence_preview_image(
                frame,
                seq.name,
                cfg,
                {},
                False,
                False,
                max_size=(800, 600),
            )

            self.assertEqual(image.size, (800, 400))
            self.assertEqual(meta["scale"], 4.0)

    def test_gui_state_builds_analyzer_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output = root / "summary.csv"
            frames = root / "frames.csv"
            settings = root / "settings.json"
            settings.write_text(
                '{"focus_guides":{"y_offset_mm":0},"peak_detect":{"difference_threshold":0.5,'
                '"roi2_extension_params":{"left":40,"right":40,"top":50,"bottom":30}}}',
                encoding="utf-8",
            )
            state = analyzer.GuiState(
                root_dir=str(root),
                output_csv=str(output),
                per_frame_csv=str(frames),
                settings_path=str(settings),
                focus_point="PointF(300, 256)",
                focus_points_csv="",
                provider_depth_mm="100",
                focus_y_offset_mm="0",
                roi2_left="10",
                roi2_right="11",
                roi2_top="12",
                roi2_bottom="13",
                difference_threshold="2.5",
                before_frame_index="1",
                after_strategy="last",
                include_selected_debug=False,
                max_sequences="5",
            )

            cfg = analyzer.config_from_gui_state(state)

            self.assertEqual(cfg.root_dir, root)
            self.assertEqual(cfg.output_csv, output)
            self.assertEqual(cfg.per_frame_csv, frames)
            self.assertEqual(cfg.focus_point, "PointF(300, 256)")
            self.assertEqual(cfg.provider_depth_mm, 100.0)
            self.assertEqual(cfg.roi2_extension_params, {"left": 10, "right": 11, "top": 12, "bottom": 13})
            self.assertEqual(cfg.difference_threshold, 2.5)
            self.assertEqual(cfg.after_strategy, "last")
            self.assertEqual(cfg.max_sequences, 5)
            self.assertEqual(cfg.roi3_extension_params, analyzer.ROI3_DEFAULT_PARAMS)
            self.assertIsNone(cfg.roi4_rect)
            self.assertIsNone(cfg.roi4_bottom_region_ratio)

    def test_gui_state_builds_multi_roi_config(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            state = analyzer.GuiState(
                root_dir=str(root),
                output_csv=str(root / "summary.csv"),
                per_frame_csv="",
                settings_path="",
                focus_point="PointF(300, 256)",
                focus_points_csv="",
                provider_depth_mm="100",
                focus_y_offset_mm="0",
                roi2_left="10",
                roi2_right="11",
                roi2_top="12",
                roi2_bottom="13",
                difference_threshold="2.5",
                before_frame_index="1",
                after_strategy="last",
                include_selected_debug=False,
                max_sequences="",
                roi3_left="21",
                roi3_right="22",
                roi3_top="23",
                roi3_bottom="24",
                roi4_x="5",
                roi4_y="6",
                roi4_width="70",
                roi4_height="80",
                roi4_bottom_region_ratio="",
            )

            cfg = analyzer.config_from_gui_state(state)

            self.assertEqual(cfg.roi3_extension_params, {"left": 21, "right": 22, "top": 23, "bottom": 24})
            self.assertEqual(cfg.roi4_rect, (5, 6, 75, 86))
            self.assertIsNone(cfg.roi4_bottom_region_ratio)

    def test_gui_state_rejects_roi4_fixed_rect_and_bottom_ratio_together(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            state = analyzer.GuiState(
                root_dir=str(root),
                output_csv=str(root / "summary.csv"),
                per_frame_csv="",
                settings_path="",
                focus_point="PointF(300, 256)",
                focus_points_csv="",
                provider_depth_mm="100",
                focus_y_offset_mm="0",
                roi2_left="10",
                roi2_right="11",
                roi2_top="12",
                roi2_bottom="13",
                difference_threshold="2.5",
                before_frame_index="1",
                after_strategy="last",
                include_selected_debug=False,
                max_sequences="",
                roi4_x="5",
                roi4_y="6",
                roi4_width="70",
                roi4_height="80",
                roi4_bottom_region_ratio="0.3",
            )

            with self.assertRaisesRegex(ValueError, "ROI4固定区域和底部高度比例不能同时填写"):
                analyzer.config_from_gui_state(state)


    def test_gui_state_maps_chinese_after_strategy_label(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            state = analyzer.GuiState(
                root_dir=str(root),
                output_csv=str(root / "summary.csv"),
                per_frame_csv="",
                settings_path="",
                focus_point=analyzer.DEFAULT_FOCUS_POINT,
                focus_points_csv="",
                provider_depth_mm="100",
                focus_y_offset_mm="0",
                roi2_left="40",
                roi2_right="40",
                roi2_top="50",
                roi2_bottom="30",
                difference_threshold="0.5",
                before_frame_index="1",
                after_strategy="最后一帧",
                include_selected_debug=False,
                max_sequences="",
            )

            cfg = analyzer.config_from_gui_state(state)

            self.assertEqual(cfg.after_strategy, "last")

    def test_gui_state_requires_focus_source(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            settings = root / "settings.json"
            settings.write_text(
                '{"focus_guides":{"y_offset_mm":0},"peak_detect":{"difference_threshold":0.5,'
                '"roi2_extension_params":{"left":40,"right":40,"top":50,"bottom":30}}}',
                encoding="utf-8",
            )
            state = analyzer.GuiState(
                root_dir=str(root),
                output_csv=str(root / "summary.csv"),
                per_frame_csv="",
                settings_path=str(settings),
                focus_point="",
                focus_points_csv="",
                provider_depth_mm="",
                focus_y_offset_mm="0",
                roi2_left="40",
                roi2_right="40",
                roi2_top="50",
                roi2_bottom="30",
                difference_threshold="0.5",
                before_frame_index="1",
                after_strategy="roi2_peak",
                include_selected_debug=False,
                max_sequences="",
            )

            with self.assertRaisesRegex(ValueError, "必须填写全局焦点或选择焦点CSV"):
                analyzer.config_from_gui_state(state)

    def test_gui_settings_change_refreshes_roi2_and_focus_preview(self):
        class FakeVar:
            def __init__(self, value):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        class FakeRoot:
            def __init__(self):
                self.after_calls = []
                self.cancelled = []

            def after(self, delay_ms, callback, *args):
                self.after_calls.append((delay_ms, callback, args))
                callback(*args)
                return "after-id"

            def after_cancel(self, after_id):
                self.cancelled.append(after_id)

        class FakeStatus:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.root = FakeRoot()
        gui.focus_point = FakeVar("PointF(100, 100)")
        gui.focus_x = FakeVar("100")
        gui.focus_y = FakeVar("100")
        gui.root_dir = FakeVar(".")
        gui.output_csv = FakeVar("summary.csv")
        gui.per_frame_csv = FakeVar("")
        gui.settings_path = FakeVar("")
        gui.focus_points_csv = FakeVar("")
        gui.provider_depth_mm = FakeVar("100")
        gui.focus_y_offset_mm = FakeVar("0")
        gui.roi2_left = FakeVar("5")
        gui.roi2_right = FakeVar("5")
        gui.roi2_top = FakeVar("5")
        gui.roi2_bottom = FakeVar("5")
        gui.roi3_left = FakeVar("5")
        gui.roi3_right = FakeVar("5")
        gui.roi3_top = FakeVar("5")
        gui.roi3_bottom = FakeVar("5")
        gui.roi4_x = FakeVar("")
        gui.roi4_y = FakeVar("")
        gui.roi4_width = FakeVar("")
        gui.roi4_height = FakeVar("")
        gui.roi4_bottom_region_ratio = FakeVar("")
        gui.roi_rect_vars = {
            roi_name: {field_name: FakeVar("") for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        gui.roi_shape_vars = {
            roi_name: FakeVar("矩形")
            for roi_name in analyzer.ROI_NAMES
        }
        gui.highlight_mode = FakeVar("基线倍数")
        gui.highlight_fixed_gray = FakeVar("93")
        gui.highlight_baseline_multiplier = FakeVar("1.15")
        gui.excel_output_dir = FakeVar("excel")
        gui.difference_threshold = FakeVar("0.5")
        gui.before_frame_index = FakeVar("1")
        gui.after_strategy = FakeVar("roi2_peak")
        gui.include_selected_debug = FakeVar(False)
        gui.max_sequences = FakeVar("")
        gui.status = FakeStatus()
        gui._preview_refresh_after_id = None
        gui._current_frame_paths = [Path("frame.png")]
        gui._step_sequence_dirs = [Path("seq")]
        gui._step_source_key = (".", False, "")
        gui._step_config_key = None

        calls = []
        gui.refresh_preview = lambda: calls.append(analyzer.config_from_gui_state(gui.current_state()))

        gui.roi2_left.set("20")
        gui._schedule_preview_refresh()
        self.assertEqual(calls[-1].roi2_extension_params["left"], 20)

        gui.focus_x.set("120")
        gui.focus_y.set("130")
        gui._schedule_preview_refresh()
        self.assertEqual(calls[-1].focus_point, "PointF(120.0, 130.0)")

    def test_gui_update_roi_stats_panel_sets_sidebar_values(self):
        class FakeVar:
            def __init__(self):
                self.value = ""

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        def fake_values():
            return {field: FakeVar() for field, _label in analyzer.ROI_STAT_DISPLAY_FIELDS}

        gui._roi_stat_vars = {
            "ROI1": fake_values(),
            "ROI2": fake_values(),
        }
        gui._update_roi_stats_panel(
            {
                "roi_stats": {
                    "ROI1": {
                        "rect": "0,0,200,100",
                        "width": "200",
                        "height": "100",
                        "area": "20000",
                        "mean": "10.000000",
                        "std": "1.500000",
                        "p90": "12.000000",
                        "threshold": "11.500000",
                        "highlight_ratio": "0.125000",
                        "hem_z_area": "25",
                        "mean_delta": "2.000000",
                    },
                    "ROI2": {"rect": "", "width": "", "height": "", "area": "", "mean": ""},
                }
            }
        )

        self.assertEqual(gui._roi_stat_vars["ROI1"]["rect"].value, "0,0,200,100")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["size"].value, "200 × 100")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["area"].value, "20000")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["mean"].value, "10")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["std"].value, "1.5")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["p90"].value, "12")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["threshold"].value, "11.5")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["highlight_ratio"].value, "0.125")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["hem_z_area"].value, "25")
        self.assertEqual(gui._roi_stat_vars["ROI1"]["mean_delta"].value, "2")
        self.assertEqual(gui._roi_stat_vars["ROI2"]["rect"].value, "-")
        self.assertEqual(gui._roi_stat_vars["ROI2"]["size"].value, "-")

    def test_gui_layout_uses_wide_two_column_roi_stats_and_left_preview_anchor(self):
        positions = [analyzer.roi_stat_card_grid_position(index) for index in range(4)]

        self.assertEqual(positions, [(0, 0), (0, 1), (1, 0), (1, 1)])
        self.assertEqual(analyzer.PREVIEW_IMAGE_ANCHOR, "nw")
        self.assertGreaterEqual(analyzer.ROI_STATS_PANEL_MIN_WIDTH, 720)
        self.assertLessEqual(analyzer.ROI_STATS_FONT_SIZE, 8)
        self.assertEqual(analyzer.ROI_STATS_VALUE_COLUMNS, 4)
        self.assertTrue(all(len(label) <= 4 for _field, label in analyzer.ROI_STAT_DISPLAY_FIELDS))

    def test_gui_stat_display_value_uses_at_most_three_decimals(self):
        self.assertEqual(analyzer._format_roi_stat_display_value("15.588854"), "15.589")
        self.assertEqual(analyzer._format_roi_stat_display_value("10.000000"), "10")
        self.assertEqual(analyzer._format_roi_stat_display_value("0.125000"), "0.125")
        self.assertEqual(analyzer._format_roi_stat_display_value("0.123400"), "0.123")
        self.assertEqual(analyzer._format_roi_stat_display_value("0,0,600,512"), "0,0,600,512")

    def test_roi_rect_overrides_replace_algorithm_rects_for_preview(self):
        frame = np.zeros((100, 120, 3), dtype=np.uint8)
        cfg = analyzer.AnalyzerConfig(
            root_dir=Path("."),
            output_csv=Path("summary.csv"),
            per_frame_csv=None,
            settings_path=None,
            focus_point="PointF(50, 50)",
            focus_points_csv=None,
            provider_depth_mm=None,
            focus_y_offset_mm=0.0,
            roi2_extension_params={"left": 5, "right": 5, "top": 5, "bottom": 5},
            difference_threshold=0.5,
            before_frame_index=1,
            after_strategy="roi2_peak",
            include_selected_debug=False,
            max_sequences=None,
            roi3_extension_params={"left": 5, "right": 5, "top": 10, "bottom": 20},
            roi_rect_overrides={
                "ROI1": (1, 2, 31, 42),
                "ROI2": (10, 11, 30, 41),
                "ROI3": (20, 21, 40, 51),
                "ROI4": (30, 31, 50, 61),
            },
        )

        meta = analyzer.resolve_roi_rects(frame, (50, 50), cfg)

        self.assertEqual(meta["roi1_rect"], (1, 2, 31, 42))
        self.assertEqual(meta["roi2_rect"], (10, 11, 30, 41))
        self.assertEqual(meta["roi3_rect"], (20, 21, 40, 51))
        self.assertEqual(meta["roi4_rect"], (30, 31, 50, 61))

    def test_save_roi_rect_overrides_persists_tool_settings(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings_path = Path(tmp) / "settings.json"
            settings_path.write_text('{"peak_detect": {"difference_threshold": 0.5}}', encoding="utf-8")
            overrides = {
                "ROI1": (1, 2, 31, 42),
                "ROI2": (10, 11, 30, 41),
                "ROI3": None,
                "ROI4": (30, 31, 50, 61),
            }

            analyzer.save_roi_rect_overrides(settings_path, overrides)
            settings = analyzer._load_settings(settings_path)

            tool_settings = settings["hem_roi2_batch_analyzer"]
            self.assertEqual(tool_settings["roi_rect_overrides"]["ROI1"], {"x": 1, "y": 2, "width": 30, "height": 40})
            self.assertEqual(tool_settings["roi_rect_overrides"]["ROI2"], {"x": 10, "y": 11, "width": 20, "height": 30})
            self.assertNotIn("ROI3", tool_settings["roi_rect_overrides"])
            self.assertEqual(analyzer._settings_roi_rect_overrides(settings)["ROI4"], (30, 31, 50, 61))

    def test_save_roi_definitions_and_highlight_rule_persist_tool_settings(self):
        with tempfile.TemporaryDirectory() as tmp:
            settings_path = Path(tmp) / "settings.json"
            settings_path.write_text('{"peak_detect": {"difference_threshold": 0.5}}', encoding="utf-8")
            definitions = {
                "ROI1": {"shape": "ellipse", "rect": (1, 2, 31, 42)},
                "ROI2": {"shape": "rectangle", "rect": (10, 11, 30, 41)},
            }
            highlight_rule = {"mode": "fixed_gray", "fixed_gray": 93, "baseline_multiplier": 1.15}

            analyzer.save_visual_analysis_settings(settings_path, definitions, highlight_rule, Path("exports"))
            settings = analyzer._load_settings(settings_path)

            self.assertEqual(
                analyzer._settings_roi_definitions(settings)["ROI1"],
                {"shape": "ellipse", "rect": (1, 2, 31, 42)},
            )
            self.assertEqual(analyzer._settings_highlight_rule(settings), highlight_rule)
            self.assertEqual(analyzer._settings_excel_output_dir(settings), Path("exports"))

    def test_roi_definitions_from_preview_lock_visible_rects_when_inputs_blank(self):
        inputs = {
            roi_name: {field_name: "" for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        shapes = {"ROI1": "ellipse", "ROI2": "rectangle", "ROI3": "ellipse", "ROI4": "rectangle"}
        preview_meta = {
            "roi1_rect": (0, 0, 100, 80),
            "roi2_rect": (20, 20, 60, 60),
            "roi3_rect": (25, 60, 65, 90),
            "roi4_rect": (0, 70, 100, 80),
        }

        definitions = analyzer._roi_definitions_from_inputs_or_preview(inputs, shapes, preview_meta)

        self.assertEqual(definitions["ROI1"], {"shape": "ellipse", "rect": (0, 0, 100, 80)})
        self.assertEqual(definitions["ROI2"], {"shape": "rectangle", "rect": (20, 20, 60, 60)})
        self.assertEqual(definitions["ROI3"], {"shape": "ellipse", "rect": (25, 60, 65, 90)})
        self.assertEqual(definitions["ROI4"], {"shape": "rectangle", "rect": (0, 70, 100, 80)})

    def test_apply_roi_definitions_to_vars_fills_saved_roi_inputs(self):
        class FakeVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.roi_rect_vars = {
            roi_name: {field_name: FakeVar("") for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        gui.roi_shape_vars = {roi_name: FakeVar("矩形") for roi_name in analyzer.ROI_NAMES}
        definitions = {
            "ROI2": {"shape": "ellipse", "rect": (20, 30, 80, 90)},
            "ROI3": {"shape": "rectangle", "rect": (25, 70, 75, 120)},
        }

        gui._apply_roi_definitions_to_vars(definitions)

        self.assertEqual(
            {field: gui.roi_rect_vars["ROI2"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "20", "y": "30", "width": "60", "height": "60"},
        )
        self.assertEqual(gui.roi_shape_vars["ROI2"].get(), "椭圆")
        self.assertEqual(
            {field: gui.roi_rect_vars["ROI3"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "25", "y": "70", "width": "50", "height": "50"},
        )
        self.assertEqual(gui.roi_shape_vars["ROI3"].get(), "矩形")

    def test_gui_syncs_current_preview_roi_rects_to_inputs(self):
        class FakeVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.roi_rect_vars = {
            roi_name: {field_name: FakeVar("") for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        gui.roi_shape_vars = {
            roi_name: FakeVar("妞渾" if roi_name == "ROI2" else "鐭╁舰")
            for roi_name in analyzer.ROI_NAMES
        }
        meta = {
            "roi1_rect": (0, 0, 600, 512),
            "roi2_rect": (259, 235, 339, 315),
            "roi3_rect": (269, 235, 329, 385),
            "roi4_rect": (0, 358, 600, 512),
        }

        gui._sync_current_preview_roi_rect_inputs(meta)

        self.assertEqual(
            {field: gui.roi_rect_vars["ROI1"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "0", "y": "0", "width": "600", "height": "512"},
        )
        self.assertEqual(
            {field: gui.roi_rect_vars["ROI2"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "259", "y": "235", "width": "80", "height": "80"},
        )
        self.assertEqual(
            {field: gui.roi_rect_vars["ROI3"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "269", "y": "235", "width": "60", "height": "150"},
        )
        self.assertEqual(
            {field: gui.roi_rect_vars["ROI4"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "0", "y": "358", "width": "600", "height": "154"},
        )
        self.assertEqual(gui.roi_shape_vars["ROI2"].get(), "妞渾")
        self.assertFalse(gui._syncing_roi_rect_inputs)

    def test_gui_roi_shape_toggle_keeps_existing_rect_parameters(self):
        class FakeVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.roi_rect_vars = {
            roi_name: {field_name: FakeVar("") for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        gui.roi_shape_vars = {roi_name: FakeVar("鐭╁舰") for roi_name in analyzer.ROI_NAMES}
        gui._current_preview_meta = {"roi2_rect": (10, 20, 40, 60)}
        gui.roi_rect_vars["ROI2"]["x"].set("10")
        gui.roi_rect_vars["ROI2"]["y"].set("20")
        gui.roi_rect_vars["ROI2"]["width"].set("30")
        gui.roi_rect_vars["ROI2"]["height"].set("40")
        calls = []
        gui._schedule_preview_refresh = lambda *_args: calls.append("refresh")

        gui.roi_shape_vars["ROI2"].set("妞渾")
        gui._on_roi_shape_changed("ROI2")
        gui.roi_shape_vars["ROI2"].set("鐭╁舰")
        gui._on_roi_shape_changed("ROI2")

        self.assertEqual(
            {field: gui.roi_rect_vars["ROI2"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "10", "y": "20", "width": "30", "height": "40"},
        )
        self.assertEqual(calls, ["refresh", "refresh"])

    def test_gui_roi_shape_toggle_fills_blank_inputs_from_preview_rect(self):
        class FakeVar:
            def __init__(self, value=""):
                self.value = value

            def get(self):
                return self.value

            def set(self, value):
                self.value = value

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.roi_rect_vars = {
            roi_name: {field_name: FakeVar("") for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        gui.roi_shape_vars = {roi_name: FakeVar("鐭╁舰") for roi_name in analyzer.ROI_NAMES}
        gui._current_preview_meta = {"roi3_rect": (25, 70, 75, 120)}
        calls = []
        gui._schedule_preview_refresh = lambda *_args: calls.append("refresh")

        gui.roi_shape_vars["ROI3"].set("妞渾")
        gui._on_roi_shape_changed("ROI3")

        self.assertEqual(
            {field: gui.roi_rect_vars["ROI3"][field].get() for field in analyzer.ROI_RECT_FIELDS},
            {"x": "25", "y": "70", "width": "50", "height": "50"},
        )
        self.assertEqual(calls, ["refresh"])

    def test_gui_schedule_preview_refresh_ignores_internal_roi_input_sync(self):
        calls = []
        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui._syncing_roi_rect_inputs = True
        gui._sync_focus_point_from_xy = lambda: calls.append("sync-focus")
        gui._current_frame_paths = [Path("frame.png")]

        gui._schedule_preview_refresh()

        self.assertEqual(calls, [])

    def test_saved_preview_roi_centers_do_not_move_after_focus_changes(self):
        frame = np.zeros((100, 120, 3), dtype=np.uint8)
        base_cfg = analyzer.AnalyzerConfig(
            root_dir=Path("."),
            output_csv=Path("summary.csv"),
            per_frame_csv=None,
            settings_path=None,
            focus_point="PointF(50, 50)",
            focus_points_csv=None,
            provider_depth_mm=None,
            focus_y_offset_mm=0.0,
            roi2_extension_params={"left": 10, "right": 10, "top": 10, "bottom": 10},
            difference_threshold=0.5,
            before_frame_index=1,
            after_strategy="roi2_peak",
            include_selected_debug=False,
            max_sequences=None,
            roi3_extension_params={"left": 10, "right": 10, "top": 20, "bottom": 40},
            roi4_bottom_region_ratio=0.3,
        )
        first_meta = analyzer.resolve_roi_rects(frame, (50, 50), base_cfg)
        blank_inputs = {
            roi_name: {field_name: "" for field_name in analyzer.ROI_RECT_FIELDS}
            for roi_name in analyzer.ROI_NAMES
        }
        shapes = {"ROI1": "rectangle", "ROI2": "ellipse", "ROI3": "rectangle", "ROI4": "ellipse"}
        definitions = analyzer._roi_definitions_from_inputs_or_preview(blank_inputs, shapes, first_meta)
        locked_cfg = analyzer.AnalyzerConfig(
            root_dir=Path("."),
            output_csv=Path("summary.csv"),
            per_frame_csv=None,
            settings_path=None,
            focus_point="PointF(80, 70)",
            focus_points_csv=None,
            provider_depth_mm=None,
            focus_y_offset_mm=0.0,
            roi2_extension_params=base_cfg.roi2_extension_params,
            difference_threshold=0.5,
            before_frame_index=1,
            after_strategy="roi2_peak",
            include_selected_debug=False,
            max_sequences=None,
            roi3_extension_params=base_cfg.roi3_extension_params,
            roi_rect_overrides=analyzer._roi_rect_overrides_from_definitions(definitions),
            roi_definitions=definitions,
        )

        changed_focus_meta = analyzer.resolve_roi_rects(frame, (80, 70), locked_cfg)

        for roi_name in analyzer.ROI_NAMES:
            key = f"{roi_name.lower()}_rect"
            self.assertEqual(changed_focus_meta[key], first_meta[key])
            self.assertEqual(changed_focus_meta["roi_shapes"][roi_name], shapes[roi_name])

    def test_export_sequence_excel_writes_roi_stats_and_histograms(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq = root / "seq_a"
            seq.mkdir()
            frame_a = np.full((20, 20, 3), 10, dtype=np.uint8)
            frame_b = np.full((20, 20, 3), 20, dtype=np.uint8)
            Image.fromarray(frame_a).save(seq / "00001_2026-06-14_00-00-00.000_frame.png")
            Image.fromarray(frame_b).save(seq / "00002_2026-06-14_00-00-00.070_frame.png")
            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=None,
                settings_path=None,
                focus_point="PointF(10, 10)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 2, "right": 2, "top": 2, "bottom": 2},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
                roi3_extension_params={"left": 2, "right": 2, "top": 2, "bottom": 2},
                roi_rect_overrides={
                    "ROI1": (0, 0, 4, 4),
                    "ROI2": (5, 5, 9, 9),
                    "ROI3": (10, 10, 14, 14),
                    "ROI4": (15, 15, 19, 19),
                },
                roi_definitions={
                    "ROI1": {"shape": "rectangle", "rect": (0, 0, 4, 4)},
                    "ROI2": {"shape": "ellipse", "rect": (5, 5, 9, 9)},
                    "ROI3": {"shape": "rectangle", "rect": (10, 10, 14, 14)},
                    "ROI4": {"shape": "rectangle", "rect": (15, 15, 19, 19)},
                },
                highlight_rule={"mode": "fixed_gray", "fixed_gray": 15, "baseline_multiplier": 1.15},
            )
            output_path = root / "seq_a.xlsx"

            analyzer.export_sequence_excel(seq, cfg, {}, output_path)

            from openpyxl import load_workbook

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            self.assertEqual(set(workbook.sheetnames), {"Summary", "Frame_ROI_Stats", "Histograms", "ROI_Config"})
            self.assertEqual(workbook["Frame_ROI_Stats"].max_row, 1 + 2 * 4)
            self.assertEqual(workbook["Histograms"].max_row, 1 + 2 * 4)
            histogram_headers = [cell.value for cell in next(workbook["Histograms"].iter_rows(min_row=1, max_row=1))]
            self.assertIn("gray_0", histogram_headers)
            self.assertIn("gray_255", histogram_headers)
            first_histogram_row = [cell.value for cell in next(workbook["Histograms"].iter_rows(min_row=2, max_row=2))]
            gray_values = first_histogram_row[4:]
            self.assertEqual(sum(value or 0 for value in gray_values), 16)
            workbook.close()

    def test_gui_default_maximize_uses_zoomed_state(self):
        class FakeRoot:
            def __init__(self):
                self.states = []

            def state(self, value):
                self.states.append(value)

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.root = FakeRoot()

        gui._maximize_root()

        self.assertEqual(gui.root.states, ["zoomed"])

    def test_gui_run_analysis_returns_immediately_while_one_sequence_runs(self):
        class FakeButton:
            def __init__(self):
                self.state_calls = []
                self.config_calls = []

            def state(self, values):
                self.state_calls.append(tuple(values))

            def configure(self, **kwargs):
                self.config_calls.append(kwargs)

        class FakeRoot:
            def __init__(self):
                self.after_calls = []

            def after(self, delay_ms, callback, *args):
                self.after_calls.append((delay_ms, callback, args))

        class FakeLabel:
            def __init__(self):
                self.configure_calls = []

            def configure(self, **kwargs):
                self.configure_calls.append(kwargs)

        class FakeScale:
            def __init__(self):
                self.configure_calls = []
                self.values = []

            def configure(self, **kwargs):
                self.configure_calls.append(kwargs)

            def set(self, value):
                self.values.append(value)

        gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
        gui.root = FakeRoot()
        gui.run_button = FakeButton()
        gui.image_label = FakeLabel()
        gui.timeline = FakeScale()
        gui.status = type("Status", (), {"set": lambda self, value: setattr(self, "value", value)})()
        gui.sequence_info = type("Status", (), {"set": lambda self, value: setattr(self, "value", value)})()
        gui.frame_info = type("Status", (), {"set": lambda self, value: setattr(self, "value", value)})()
        gui.show_roi1 = type("Var", (), {"get": lambda self: True})()
        gui.show_roi2 = type("Var", (), {"get": lambda self: True})()
        gui.show_roi3 = type("Var", (), {"get": lambda self: False})()
        gui.show_roi4 = type("Var", (), {"get": lambda self: False})()
        gui.show_focus = type("Var", (), {"get": lambda self: True})()
        gui._analysis_running = False
        gui._step_config_key = None
        gui._step_source_key = None
        gui._step_sequence_dirs = []
        gui._step_next_index = 0
        gui._current_sequence_index = 0
        gui._step_summary_rows = []
        gui._step_frame_rows = []
        gui._analyzed_sequences = set()
        gui._current_frame_paths = []
        gui._current_frame_index = 0
        gui._current_preview_image = None
        gui._current_preview_meta = {}
        gui._photo_image = None
        gui._display_preview_image = lambda image: setattr(gui, "_current_preview_image", image)
        gui.current_state = lambda: analyzer.GuiState(
            root_dir=".",
            output_csv="summary.csv",
            per_frame_csv="",
            settings_path="",
            focus_point=analyzer.DEFAULT_FOCUS_POINT,
            focus_points_csv="",
            provider_depth_mm="",
            focus_y_offset_mm="0",
            roi2_left="40",
            roi2_right="40",
            roi2_top="50",
            roi2_bottom="30",
            difference_threshold="0.5",
            before_frame_index="1",
            after_strategy="roi2_peak",
            include_selected_debug=False,
            max_sequences="",
        )

        calls = []
        worker_started = threading.Event()
        release_worker = threading.Event()

        def slow_analyze_sequence(sequence_dir, config, focus_points):
            calls.append(sequence_dir.name)
            worker_started.set()
            release_worker.wait(0.5)
            return (
                {
                    "sequence": sequence_dir.name,
                    "status": "ok",
                    "error": "",
                    "frame_count": "0",
                    "focus_anchor": "",
                    "offset_anchor": "",
                    "roi2_rect": "",
                    "before_frame_index": "",
                    "before_frame": "",
                    "before_mean": "",
                    "after_frame_index": "",
                    "after_frame": "",
                    "after_mean": "",
                    "roi2_diff": "",
                    "difference_threshold": "",
                    "roi2_color": "red",
                    "after_strategy": "",
                },
                [],
            )

        def fail_analyze_root(config, progress_callback=None):
            raise AssertionError("GUI must not batch-analyze all sequences")

        original_sequence = analyzer.analyze_sequence
        original = analyzer.analyze_root
        import tkinter.messagebox as messagebox

        original_showinfo = messagebox.showinfo
        original_showerror = messagebox.showerror
        analyzer.analyze_sequence = slow_analyze_sequence
        analyzer.analyze_root = fail_analyze_root
        messagebox.showinfo = lambda *args, **kwargs: None
        messagebox.showerror = lambda *args, **kwargs: None
        try:
            with tempfile.TemporaryDirectory() as tmp:
                root = Path(tmp)
                seq_a = root / "seq_a"
                seq_b = root / "seq_b"
                seq_a.mkdir()
                seq_b.mkdir()
                write_frame(seq_a / "00001_2026-06-14_00-00-00.000_frame.png", 10, size=(400, 400))
                write_frame(seq_b / "00001_2026-06-14_00-00-00.000_frame.png", 10, size=(400, 400))
                gui.current_state = lambda: analyzer.GuiState(
                    root_dir=str(root),
                    output_csv=str(root / "summary.csv"),
                    per_frame_csv="",
                    settings_path="",
                    focus_point=analyzer.DEFAULT_FOCUS_POINT,
                    focus_points_csv="",
                    provider_depth_mm="100",
                    focus_y_offset_mm="0",
                    roi2_left="40",
                    roi2_right="40",
                    roi2_top="50",
                    roi2_bottom="30",
                    difference_threshold="0.5",
                    before_frame_index="1",
                    after_strategy="roi2_peak",
                    include_selected_debug=False,
                    max_sequences="",
                )
                started = time.perf_counter()
                gui.run_analysis()
                elapsed = time.perf_counter() - started
                self.assertTrue(worker_started.wait(1.0))
        finally:
            release_worker.set()
            analyzer.analyze_sequence = original_sequence
            analyzer.analyze_root = original
            messagebox.showinfo = original_showinfo
            messagebox.showerror = original_showerror

        self.assertLess(elapsed, 0.1)
        self.assertEqual(gui.run_button.state_calls, [("disabled",)])
        self.assertTrue(gui._analysis_running)
        self.assertEqual(calls, ["seq_a"])

    def test_gui_stepwise_analysis_advances_one_sequence_per_click(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            seq_a = root / "seq_a"
            seq_b = root / "seq_b"
            seq_a.mkdir()
            seq_b.mkdir()
            write_frame(seq_a / "00001_2026-06-14_00-00-00.000_frame.png", 10, size=(400, 400))
            write_frame(seq_a / "00002_2026-06-14_00-00-00.070_frame.png", 20, size=(400, 400))
            write_frame(seq_b / "00001_2026-06-14_00-00-00.000_frame.png", 10, size=(400, 400))
            write_frame(seq_b / "00002_2026-06-14_00-00-00.070_frame.png", 12, size=(400, 400))

            cfg = analyzer.AnalyzerConfig(
                root_dir=root,
                output_csv=root / "summary.csv",
                per_frame_csv=root / "frames.csv",
                settings_path=None,
                focus_point="PointF(100, 100)",
                focus_points_csv=None,
                provider_depth_mm=100.0,
                focus_y_offset_mm=0.0,
                roi2_extension_params={"left": 2, "right": 2, "top": 3, "bottom": 3},
                difference_threshold=5.0,
                before_frame_index=1,
                after_strategy="last",
                include_selected_debug=False,
                max_sequences=None,
            )
            gui = object.__new__(analyzer.HemRoi2BatchAnalyzerGui)
            gui._step_config_key = None
            gui._step_sequence_dirs = []
            gui._step_next_index = 0
            gui._step_summary_rows = []
            gui._step_frame_rows = []

            gui._reset_step_state(cfg)
            focus_points = analyzer.load_focus_points_csv(cfg.focus_points_csv)
            first_row, first_frames = analyzer.analyze_sequence(gui._step_sequence_dirs[gui._step_next_index], cfg, focus_points)
            gui._step_summary_rows.append(first_row)
            gui._step_frame_rows.extend(first_frames)
            gui._step_next_index += 1
            analyzer.write_csv(cfg.output_csv, analyzer.SUMMARY_FIELDS, gui._step_summary_rows)
            analyzer.write_csv(cfg.per_frame_csv, analyzer.FRAME_FIELDS, gui._step_frame_rows)

            with cfg.output_csv.open("r", encoding="utf-8-sig", newline="") as f:
                saved = list(csv.DictReader(f))
            self.assertEqual([row["sequence"] for row in saved], ["seq_a"])

            second_row, second_frames = analyzer.analyze_sequence(gui._step_sequence_dirs[gui._step_next_index], cfg, focus_points)
            gui._step_summary_rows.append(second_row)
            gui._step_frame_rows.extend(second_frames)
            gui._step_next_index += 1
            analyzer.write_csv(cfg.output_csv, analyzer.SUMMARY_FIELDS, gui._step_summary_rows)
            analyzer.write_csv(cfg.per_frame_csv, analyzer.FRAME_FIELDS, gui._step_frame_rows)

            with cfg.output_csv.open("r", encoding="utf-8-sig", newline="") as f:
                saved = list(csv.DictReader(f))
            self.assertEqual([row["sequence"] for row in saved], ["seq_a", "seq_b"])
            self.assertEqual(gui._step_next_index, 2)


if __name__ == "__main__":
    unittest.main()
